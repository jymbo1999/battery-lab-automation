"""Minimal XLSX reader for EIS files whose Z' and Z'' columns are known."""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Tuple

import numpy as np
from openpyxl import load_workbook


def read_xlsx_z_columns(
    path: str | Path,
    *,
    sheet_name: Optional[str] = None,
    z_real_col: int = 3,
    z_imag_col: int = 4,
    first_data_row: int = 2,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Reads Z' and Z'' from an Excel file.

    Defaults match the user's current EIS XLSX convention:
        C column = Z'
        D column = Z''
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    z_real = []
    z_imag = []
    for row in range(first_data_row, ws.max_row + 1):
        xr = ws.cell(row=row, column=z_real_col).value
        yi = ws.cell(row=row, column=z_imag_col).value
        if xr is None or yi is None:
            continue
        try:
            z_real.append(float(xr))
            z_imag.append(float(yi))
        except (TypeError, ValueError):
            continue

    return np.asarray(z_real, dtype=float), np.asarray(z_imag, dtype=float)
