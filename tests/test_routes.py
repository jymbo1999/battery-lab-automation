import json
import tempfile
import unittest
from contextlib import ExitStack
from pathlib import Path
from unittest.mock import patch

from flask import Flask
from openpyxl import Workbook, load_workbook

from battery_lab import register_battery_lab
from battery_lab import routes


class BatteryRouteTests(unittest.TestCase):
    def make_client(self, root: Path):
        app = Flask(__name__)
        app.secret_key = "test"
        register_battery_lab(app)
        patches = [
            patch.object(routes, "BATTERY_DATA_ROOT", root),
            patch.object(routes, "BATTERY_EIS_ROOT", root / "EIS"),
            patch.object(routes, "BATTERY_CAPACITY_ROOT", root / "capacity"),
            patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"),
            patch.object(routes, "BATTERY_CONDITION_WORKBOOK", root / "Project_Abstract" / "Cell condition Calculation.xlsx"),
            patch.object(routes, "BATTERY_MATCH_EIS_JSON", root / "battery_visual_outputs" / "eis_match_overrides.json"),
            patch.object(routes, "BATTERY_MATCH_CAPACITY_JSON", root / "battery_visual_outputs" / "capacity_match_overrides.json"),
            patch.object(routes, "BATTERY_STREAMLIT_URL", ""),
        ]
        return app.test_client(), patches

    def patched_client(self, root: Path):
        client, patches = self.make_client(root)
        stack = ExitStack()
        for item in patches:
            stack.enter_context(item)
        stack.enter_context(client)
        return client, stack

    def test_named_routes_render_and_graph_routes_select_artifacts(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            (root / "EIS").mkdir(parents=True)
            (root / "capacity").mkdir(parents=True)
            (root / "Project_Abstract").mkdir(parents=True)
            eis_dir = root / "battery_visual_outputs" / "eis"
            capacity_dir = root / "battery_visual_outputs" / "capacity"
            eis_dir.mkdir(parents=True)
            capacity_dir.mkdir(parents=True)
            (eis_dir / "b.svg").write_text("<svg/>", encoding="utf-8")
            (capacity_dir / "c.svg").write_text("<svg/>", encoding="utf-8")

            client, patches = self.make_client(root)
            with patches[0], patches[1], patches[2], patches[3], patches[4], patches[5], patches[6], patches[7]:
                for path in [
                    "/battery/",
                    "/battery/journal",
                    "/battery/eis",
                    "/battery/capacity",
                    "/battery/review_EIS_capacity",
                    "/battery/files",
                    "/battery/settings",
                    "/battery/status",
                    "/battery/health",
                    "/battery/api/eis/finder",
                    "/battery/api/capacity/finder",
                    "/battery/api/capacity/viewer/source?key=missing.wrd",
                ]:
                    response = client.get(path)
                    self.assertEqual(response.status_code, 200, path)

                jobs_response = client.get("/battery/jobs")
                self.assertEqual(jobs_response.status_code, 302)
                self.assertEqual(jobs_response.headers["Location"], "/battery/eis")

                eis_response = client.get("/battery/eis?graph=b.svg")
                eis_html = eis_response.get_data(as_text=True)
                self.assertIn("Live Viewer", eis_html)
                self.assertIn("Rct</b>", eis_html)
                self.assertNotIn("Streamlit 원본처럼", eis_html)
                # 그래프 산출물 생성/갱신 job launcher moved to the Settings page
                self.assertNotIn("그래프 산출물 생성/갱신", eis_html)
                self.assertNotIn("최근 작업 새로고침", eis_html)
                self.assertNotIn("아래에서 실행과 상태 확인을 함께 처리합니다.", eis_html)
                self.assertNotIn('href="/battery/jobs"', eis_html)
                self.assertNotIn("원본 EIS 데이터 브라우저", eis_html)
                self.assertNotIn("/battery/api/eis/finder", eis_html)
                self.assertNotIn("EIS 수동 매칭", eis_html)
                self.assertNotIn("EIS 산출 그래프 보기", eis_html)
                self.assertNotIn("이미 생성된 SVG/PNG artifact를 개별 파일 단위로 확인합니다.", eis_html)

                capacity_response = client.get("/battery/capacity?graph=c.svg")
                capacity_html = capacity_response.get_data(as_text=True)
                self.assertIn("1) 0.1C continuous", capacity_html)
                self.assertIn("2) 안정화후 0.5C", capacity_html)
                self.assertIn("3) rate performance", capacity_html)
                self.assertNotIn("Protocol cluster</option>", capacity_html)
                self.assertIn("WRD/raw source preview", capacity_html)
                self.assertIn("Live Viewer", capacity_html)
                self.assertIn("rate performance", capacity_html)
                self.assertIn("R2/0.1", capacity_html)
                self.assertNotIn("protocol cluster 또는 단일", capacity_html)
                # 그래프 산출물 생성/갱신 job launcher moved to the Settings page
                self.assertNotIn("최근 작업 새로고침", capacity_html)
                self.assertNotIn('href="/battery/jobs"', capacity_html)
                self.assertNotIn("원본 Capacity 데이터 브라우저", capacity_html)
                self.assertNotIn("/battery/api/capacity/finder", capacity_html)
                self.assertNotIn("Capacity 수동 매칭", capacity_html)
                self.assertNotIn("Capacity 산출 그래프 보기", capacity_html)
                self.assertNotIn("/battery/artifact/capacity/c.svg", capacity_html)

                settings_response = client.get("/battery/settings")
                settings_html = settings_response.get_data(as_text=True)
                self.assertIn("운영 도구", settings_html)
                # job launcher now lives on Settings for both analysis types
                self.assertIn("그래프 산출물 생성/갱신 — EIS", settings_html)
                self.assertIn("그래프 산출물 생성/갱신 — Capacity", settings_html)
                self.assertIn("최근 작업 새로고침", settings_html)
                self.assertIn("/battery/files", settings_html)
                self.assertIn("/battery/review_EIS_capacity", settings_html)
                self.assertIn("/battery/api/eis/finder", settings_html)
                self.assertIn("/battery/api/capacity/finder", settings_html)
                self.assertIn("Capacity CSV / WRD Audit", settings_html)
                self.assertIn("/battery/api/capacity/csv-wrd-audit", settings_html)
                audit_response = client.post("/battery/api/capacity/csv-wrd-audit")
                self.assertEqual(audit_response.status_code, 200)
                self.assertTrue(audit_response.get_json()["ok"])

                review_response = client.get("/battery/review_EIS_capacity")
                review_html = review_response.get_data(as_text=True)
                self.assertIn("EIS / Capacity 수동 매칭 검토", review_html)
                self.assertIn("/battery/api/eis/match-review", review_html)
                self.assertIn("/battery/api/capacity/match-review", review_html)

    def test_legacy_query_tab_redirects_to_named_route(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            client, patches = self.make_client(root)
            with patches[0], patches[1], patches[2], patches[3], patches[4], patches[5], patches[6], patches[7]:
                response = client.get("/battery/?tab=eis&eis=sample.svg")

            self.assertEqual(response.status_code, 302)
            self.assertEqual(response.headers["Location"], "/battery/eis?graph=sample.svg")

    def test_journal_excel_routes_read_and_save_workbook_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            condition_path = root / "Project_Abstract" / "Cell condition Calculation.xlsx"
            condition_path.parent.mkdir(parents=True)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "JYJ"
            worksheet.append(["Sample", "date", "Binder"])
            worksheet.append(["before", "260101", "2wt%cmc"])
            workbook.save(condition_path)

            client, stack = self.patched_client(root)
            with stack:
                journal = client.get("/battery/journal").get_data(as_text=True)
                self.assertIn("/battery/journal/excel", journal)
                self.assertIn("battery-page-journal", journal)
                self.assertIn("battery-panel journal-panel", journal)
                self.assertIn("battery-panel battery-path-status", journal)
                self.assertIn(".battery-page-journal .battery-main", journal)
                self.assertIn("새 실험 등록", journal)
                self.assertIn("/battery/api/import/drafts", journal)
                self.assertIn("파일 업로드", journal)
                self.assertIn("실험정보 입력하기", journal)
                self.assertIn("미리보기", journal)
                self.assertIn("실험일지 xlsx에 저장하기", journal)
                self.assertIn("/commit", journal)

                excel = client.get("/battery/journal/excel").get_data(as_text=True)
                self.assertIn("/battery/api/journal/sheet", excel)
                self.assertIn("/battery/api/journal/cell", excel)

                sheet = client.get("/battery/api/journal/sheet").get_json()
                self.assertEqual(sheet["sheet"], "JYJ")
                self.assertEqual(sheet["rows"][1]["cells"][0]["value"], "before")

                response = client.post("/battery/api/journal/cell", json={"row": 2, "column": 1, "value": "after"})
                self.assertEqual(response.status_code, 200)
                self.assertTrue(response.get_json()["ok"])

            reloaded = load_workbook(condition_path, data_only=False)
            self.assertEqual(reloaded["JYJ"].cell(row=2, column=1).value, "after")
            reloaded.close()

    def test_journal_sheet_hide_filter_omits_ignored_rows_for_fast_initial_load(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            condition_path = root / "Project_Abstract" / "Cell condition Calculation.xlsx"
            condition_path.parent.mkdir(parents=True)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "JYJ"
            worksheet.append(["Sample", "참고", "전해질", "종류", "Binder", "Voltage range"])
            worksheet.append(["matched", "12파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"])
            worksheet.append(["ignored", "other", "1.0M LiPF6 EC/DEC 1:1", "LIB", "PVDF", "0.01~2V"])
            workbook.save(condition_path)

            client, stack = self.patched_client(root)
            with stack:
                hidden = client.get("/battery/api/journal/sheet?filter=hide").get_json()
                full = client.get("/battery/api/journal/sheet?filter=all").get_json()

            self.assertFalse(hidden["includeIgnoredRows"])
            self.assertTrue(full["includeIgnoredRows"])
            self.assertEqual(hidden["filter"]["ignoredRows"], 1)
            self.assertEqual(full["filter"]["ignoredRows"], 1)
            hidden_row_3 = [row for row in hidden["rows"] if row["index"] == 3]
            self.assertEqual(len(hidden_row_3), 1)
            self.assertTrue(hidden_row_3[0]["extra"])
            self.assertEqual(hidden_row_3[0]["cells"][0]["value"], "")
            self.assertIn(3, {row["index"] for row in full["rows"]})

    def test_import_metadata_options_api_reads_existing_conditions(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            condition_path = root / "Project_Abstract" / "conditions.csv"
            condition_path.parent.mkdir(parents=True)
            condition_path.write_text(
                "Sample,date,종류,전해질,Binder,Voltage range,ratio,Areal mass density\n"
                "sample 1,260627,LIB,1.0M LiPF6,CMC/SBR,0.01~2V,0.95,7.1\n",
                encoding="utf-8",
            )
            client, stack = self.patched_client(root)
            with stack, patch.object(routes, "BATTERY_CONDITION_WORKBOOK", condition_path):
                response = client.get("/battery/api/import/metadata-options")

            self.assertEqual(response.status_code, 200)
            payload = response.get_json()
            self.assertTrue(payload["ok"])
            self.assertIn("date", payload["required_fields"])
            self.assertIn("sample 1", payload["options"]["sample"])
            self.assertIn("1.0M LiPF6", payload["options"]["electrolyte"])

    def test_jobs_api_reports_unavailable_without_main_app_db(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            client, stack = self.patched_client(root)
            with stack:
                response = client.get("/battery/api/jobs")

            self.assertEqual(response.status_code, 503)
            self.assertFalse(response.get_json()["available"])

    def test_ai_status_api_reports_policy_without_main_app_db(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            client, stack = self.patched_client(root)
            with stack:
                response = client.get("/battery/api/ai/status")

            self.assertEqual(response.status_code, 200)
            payload = response.get_json()
            self.assertFalse(payload["available"])
            self.assertIn("policy", payload)
            self.assertEqual(payload["policy"]["default_call_mode"], "dry_run")

    def test_eis_match_api_saves_streamlit_compatible_override_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            eis_file = root / "EIS" / "260521" / "0hr" / "1.5act 3T_01.SEO"
            eis_file.parent.mkdir(parents=True)
            eis_file.write_text("placeholder", encoding="utf-8")
            condition_path = root / "Project_Abstract" / "conditions.csv"
            condition_path.parent.mkdir(parents=True)
            condition_path.write_text(
                "Sample,date,Binder,Voltage range,ratio,Areal mass density\n"
                "1.5 act 3T_1,260521,CMC/SBR,0.01~2V,0.95,7.1\n"
                "1.5 act 3T_2,260521,CMC/SBR,0.01~2V,0.95,7.1\n",
                encoding="utf-8",
            )
            override_path = root / "battery_visual_outputs" / "eis_match_overrides.json"

            client, stack = self.patched_client(root)
            with stack, patch.object(routes, "BATTERY_CONDITION_WORKBOOK", condition_path), patch.object(routes, "BATTERY_MATCH_EIS_JSON", override_path):
                payload = client.get("/battery/api/eis/matches").get_json()
                self.assertTrue(payload["report_ready"])
                self.assertGreaterEqual(len(payload["rows"]), 2)
                selected = next(row for row in payload["rows"] if row["condition_key"] == "1.5 act 3T_2")

                response = client.post("/battery/api/eis/matches", json={"selections": [selected]})
                self.assertEqual(response.status_code, 200)

                data = json.loads(override_path.read_text(encoding="utf-8"))
                self.assertEqual(data["260521/0hr/1.5act 3T_01.SEO"]["condition_key"], "1.5 act 3T_2")
                self.assertEqual(data["260521/0hr/1.5act 3T_01.SEO"]["sample"], "1.5 act 3T_2")
                self.assertIn("selected_at", data["260521/0hr/1.5act 3T_01.SEO"])

                clear_response = client.delete("/battery/api/eis/matches")
                self.assertEqual(clear_response.status_code, 200)
                self.assertEqual(json.loads(override_path.read_text(encoding="utf-8")), {})

    def test_capacity_match_api_saves_override_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            capacity_file = root / "capacity" / "260522" / "1.5act_3T_0.5C_024_Capacity.csv"
            capacity_file.parent.mkdir(parents=True)
            capacity_file.write_text("placeholder", encoding="utf-8")
            condition_path = root / "Project_Abstract" / "conditions.csv"
            condition_path.parent.mkdir(parents=True)
            condition_path.write_text(
                "Sample,date,Areal mass density\n"
                "1.5 act 3T,260522,7.1\n"
                "1.5act 3T,260522,7.1\n",
                encoding="utf-8",
            )
            override_path = root / "battery_visual_outputs" / "capacity_match_overrides.json"

            client, stack = self.patched_client(root)
            with stack, patch.object(routes, "BATTERY_CONDITION_WORKBOOK", condition_path), patch.object(routes, "BATTERY_MATCH_CAPACITY_JSON", override_path):
                payload = client.get("/battery/api/capacity/matches").get_json()
                self.assertTrue(payload["report_ready"])
                selected = next(row for row in payload["rows"] if row["condition_key"] == "1.5 act 3T")

                response = client.post("/battery/api/capacity/matches", json={"selections": [selected]})
                self.assertEqual(response.status_code, 200)

                data = json.loads(override_path.read_text(encoding="utf-8"))
                self.assertEqual(data["260522/1.5act_3T_0.5C_024_Capacity.csv"]["condition_key"], "1.5 act 3T")

    def test_match_review_api_saves_direct_rows_and_delete_candidates(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            eis_file = root / "EIS" / "260521" / "0hr" / "1.5act 3T_01.SEO"
            bad_file = root / "EIS" / "260521" / "0hr" / "bad_file.SEO"
            eis_file.parent.mkdir(parents=True)
            eis_file.write_text("placeholder", encoding="utf-8")
            bad_file.write_text("placeholder", encoding="utf-8")
            condition_path = root / "Project_Abstract" / "conditions.csv"
            condition_path.parent.mkdir(parents=True)
            condition_path.write_text(
                "Sample,date,Binder,Voltage range,ratio,Areal mass density\n"
                "1.5 act 3T_1,260521,CMC/SBR,0.01~2V,0.95,7.1\n"
                "1.5 act 3T_2,260521,CMC/SBR,0.01~2V,0.95,7.1\n",
                encoding="utf-8",
            )
            override_path = root / "battery_visual_outputs" / "eis_match_overrides.json"

            client, stack = self.patched_client(root)
            with stack, patch.object(routes, "BATTERY_CONDITION_WORKBOOK", condition_path), patch.object(routes, "BATTERY_MATCH_EIS_JSON", override_path):
                payload = client.get("/battery/api/eis/matches").get_json()
                self.assertTrue(payload["report_ready"])
                self.assertIn("260521/0hr/1.5act 3T_01.SEO", {row["relative_path"] for row in payload["final_rows"]})

                response = client.post(
                    "/battery/api/eis/match-review",
                    json={
                        "direct_matches": [{"file": "260521/0hr/1.5act 3T_01.SEO", "journal_row": 3}],
                        "delete_files": [{"file": "260521/0hr/bad_file.SEO", "reason": "wrong_source"}],
                    },
                )
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.get_json()["saved_count"], 2)

                data = json.loads(override_path.read_text(encoding="utf-8"))
                self.assertEqual(data["260521/0hr/1.5act 3T_01.SEO"]["condition_key"], "1.5 act 3T_2")
                self.assertEqual(data["260521/0hr/1.5act 3T_01.SEO"]["journal_row"], 3)
                self.assertEqual(data["260521/0hr/1.5act 3T_01.SEO"]["selection_source"], "review_direct_row")
                self.assertTrue(data["260521/0hr/bad_file.SEO"]["delete_candidate"])
                self.assertEqual(data["260521/0hr/bad_file.SEO"]["action"], "delete_file")


if __name__ == "__main__":
    unittest.main()
