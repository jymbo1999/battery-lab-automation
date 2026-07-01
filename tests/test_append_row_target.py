"""append_journal_row must land right after the last *populated* row, not after
openpyxl's max_row (which counts blank styled/trailing rows)."""
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from battery_lab.experiment_import import append_journal_row, last_populated_row


class AppendRowTargetTests(unittest.TestCase):
    def _workbook(self, tmp: Path) -> Path:
        path = tmp / "cond.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "JYJ"
        ws.append(["참고", "전해질", "종류", "Date", "Sample"])  # row 1 headers
        ws.cell(row=2, column=5).value = "cell A"  # last real data on row 2
        # Simulate phantom trailing rows: styling on a far row inflates max_row
        # without adding any value.
        ws.cell(row=200, column=5).fill  # touch a distant cell's style
        ws.cell(row=200, column=5).number_format = "General"
        wb.save(path)
        wb.close()
        return path

    def test_last_populated_row_ignores_blank_trailing_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = self._workbook(Path(tmp))
            wb = load_workbook(path)
            ws = wb["JYJ"]
            try:
                self.assertGreaterEqual(ws.max_row, 2)
                self.assertEqual(last_populated_row(ws), 2)
            finally:
                wb.close()

    def test_append_lands_right_after_last_data(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = self._workbook(Path(tmp))
            row = append_journal_row(path, "JYJ", {"sample": "cell B"})
            self.assertEqual(row, 3)  # after row 2, not after max_row


if __name__ == "__main__":
    unittest.main()
