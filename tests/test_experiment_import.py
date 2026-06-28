import io
import json
import struct
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from flask import Flask
from openpyxl import Workbook, load_workbook

from battery_lab import register_battery_lab
from battery_lab import routes
from battery_lab.experiment_import import create_import_draft


def _pack_wrd_record(
    *,
    test_time_s=0.0,
    cycle_index=0,
    voltage=3.7,
    current=0.1,
    charge_q_ah=0.0,
    discharge_q_ah=0.0,
):
    data = bytearray()
    data += struct.pack("<q", 638000000000000000)
    data += struct.pack("<i", 1)
    data += struct.pack("<q", int(test_time_s * 1e7))
    data += struct.pack("<q", int(test_time_s * 1e7))
    data += struct.pack("<q", int(test_time_s * 1e7))
    data += struct.pack("<i", 1)
    data += struct.pack("<i", 1)
    data += struct.pack("<i", cycle_index)
    data += bytes([1, 1, 1])
    data += struct.pack("<i", 0)
    current_range = b"101mA"
    data += bytes([len(current_range)])
    data += current_range
    values = [
        voltage,
        current,
        charge_q_ah,
        discharge_q_ah,
        charge_q_ah * voltage,
        discharge_q_ah * voltage,
        0.0,
        25.0,
        voltage,
    ]
    data += struct.pack("<9d", *values)
    return bytes(data)


def synthetic_wrd_bytes() -> bytes:
    header = b"metadata DATE TIME VOLTAGE CURRENT CHARGE Q DISCHARGE Q" + b"\x00" * 100
    return header + b"".join(
        [
            _pack_wrd_record(test_time_s=0, cycle_index=0, charge_q_ah=0.001, discharge_q_ah=0.0),
            _pack_wrd_record(test_time_s=10, cycle_index=0, charge_q_ah=0.002, discharge_q_ah=0.0018),
            _pack_wrd_record(test_time_s=20, cycle_index=1, charge_q_ah=0.001, discharge_q_ah=0.0),
            _pack_wrd_record(test_time_s=30, cycle_index=1, charge_q_ah=0.003, discharge_q_ah=0.0024),
            _pack_wrd_record(test_time_s=40, cycle_index=2, charge_q_ah=0.001, discharge_q_ah=0.0),
            _pack_wrd_record(test_time_s=50, cycle_index=2, charge_q_ah=0.0031, discharge_q_ah=0.0025),
        ]
    )


class ExperimentImportDraftTests(unittest.TestCase):
    def test_create_import_draft_parses_wrd_and_writes_preview_manifest(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_root = Path(tmp) / "battery_visual_outputs"

            manifest = create_import_draft(
                [("sample.wrd", io.BytesIO(synthetic_wrd_bytes()))],
                output_root,
                draft_id="draft-test",
                write_raw_wrd=True,
            )

            self.assertEqual(manifest.draft_id, "draft-test")
            self.assertEqual(manifest.errors, [])
            self.assertEqual(len(manifest.files), 1)
            item = manifest.files[0]
            self.assertEqual(item.parser_kind, "wrd")
            self.assertEqual(item.analysis_type, "capacity")
            self.assertEqual(item.normalized_rows, 3)
            self.assertTrue(Path(item.processed_path).exists())
            self.assertTrue(Path(item.parser_meta_path).exists())
            self.assertTrue(Path(item.raw_timeseries_path).exists())
            self.assertTrue(Path(item.plot_path).exists())
            self.assertTrue(Path(item.plot_meta_path).exists())
            self.assertEqual(item.metrics["ce_formula"], "charge_over_discharge")
            self.assertEqual(item.metrics["ice_percent"], 111.111)
            self.assertEqual(item.suggested_assignment, "capacity_1")
            self.assertEqual(item.assignment, "capacity_1")
            self.assertIn("capacity_3", item.assignment_options)
            self.assertTrue((output_root / "import_drafts" / "draft-test" / "manifest.json").exists())

    def test_import_draft_api_accepts_multipart_upload(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            app = Flask(__name__)
            app.secret_key = "test"
            register_battery_lab(app)
            client = app.test_client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"):
                response = client.post(
                    "/battery/api/import/drafts",
                    data={"files": (io.BytesIO(synthetic_wrd_bytes()), "sample.wrd")},
                    content_type="multipart/form-data",
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                self.assertTrue(payload["ok"])
                self.assertEqual(payload["files"][0]["parser_kind"], "wrd")
                self.assertEqual(payload["files"][0]["analysis_type"], "capacity")
                self.assertEqual(payload["files"][0]["metrics"]["ice_percent"], 111.111)
                self.assertEqual(payload["files"][0]["assignment"], "capacity_1")
                self.assertEqual(payload["files"][0]["suggested_assignment"], "capacity_1")
                self.assertIn("/battery/api/import/drafts/", payload["files"][0]["plot_url"])
                artifact = client.get(payload["files"][0]["plot_url"])
                assignment_response = client.patch(
                    f"/battery/api/import/drafts/{payload['draft_id']}/assignments",
                    json={"assignments": {payload["files"][0]["file_id"]: "capacity_3"}},
                )
                metadata_response = client.patch(
                    f"/battery/api/import/drafts/{payload['draft_id']}/metadata",
                    json={
                        "metadata": {
                            "date": "260627",
                            "sample": "sample 1",
                            "cell_type": "LIB",
                            "electrolyte": "1.0M LiPF6",
                            "binder": "CMC/SBR",
                            "voltage_range": "0.01~2V",
                            "ratio": "0.95",
                            "areal_mass_density": "7.1",
                        }
                    },
                )
                condition_path = root / "Project_Abstract" / "conditions.csv"
                condition_path.parent.mkdir(parents=True)
                condition_path.write_text(
                    "Sample,date,종류,전해질,Binder,Voltage range,ratio,Areal mass density\n"
                    "sample existing,260626,LIB,1.0M LiPF6,CMC/SBR,0.01~2V,0.95,7.1\n",
                    encoding="utf-8",
                )
                with patch.object(routes, "BATTERY_CONDITION_WORKBOOK", condition_path):
                    cluster_response = client.get(f"/battery/api/import/drafts/{payload['draft_id']}/cluster-preview")
                workbook_path = root / "Project_Abstract" / "Cell condition Calculation.xlsx"
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "JYJ"
                worksheet.append(["Sample", "date", "종류", "전해질", "Binder", "Voltage range", "ratio", "Areal mass density"])
                workbook.save(workbook_path)
                workbook.close()
                capacity_override_path = root / "battery_visual_outputs" / "capacity_match_overrides.json"
                eis_override_path = root / "battery_visual_outputs" / "eis_match_overrides.json"
                with (
                    patch.object(routes, "BATTERY_CONDITION_WORKBOOK", workbook_path),
                    patch.object(routes, "BATTERY_CAPACITY_ROOT", root / "capacity"),
                    patch.object(routes, "BATTERY_EIS_ROOT", root / "EIS"),
                    patch.object(routes, "BATTERY_MATCH_CAPACITY_JSON", capacity_override_path),
                    patch.object(routes, "BATTERY_MATCH_EIS_JSON", eis_override_path),
                ):
                    commit_response = client.post(f"/battery/api/import/drafts/{payload['draft_id']}/commit")
                    match_refresh_response = client.get("/battery/api/capacity/matches")

            self.assertEqual(artifact.status_code, 200)
            self.assertIn(b"<svg", artifact.get_data())
            self.assertEqual(assignment_response.status_code, 200)
            assignment_payload = assignment_response.get_json()
            self.assertTrue(assignment_payload["ok"])
            self.assertEqual(assignment_payload["files"][0]["assignment"], "capacity_3")
            self.assertEqual(metadata_response.status_code, 200)
            metadata_payload = metadata_response.get_json()
            self.assertTrue(metadata_payload["ok"])
            self.assertEqual(metadata_payload["metadata_status"], "ready")
            self.assertEqual(metadata_payload["metadata"]["date"], "260627")
            self.assertEqual(cluster_response.status_code, 200)
            cluster_payload = cluster_response.get_json()
            self.assertTrue(cluster_payload["ok"])
            self.assertEqual(cluster_payload["rows"][0]["status"], "matched_existing_cluster")
            self.assertEqual(cluster_payload["rows"][0]["existing_match_count"], 1)
            self.assertEqual(commit_response.status_code, 200)
            commit_payload = commit_response.get_json()
            self.assertTrue(commit_payload["ok"])
            self.assertEqual(commit_payload["commit_status"], "committed")
            self.assertEqual(commit_payload["journal_row"], 2)
            self.assertEqual(len(commit_payload["saved_files"]), 1)
            self.assertEqual(commit_payload["queued_jobs"], [])
            saved = commit_payload["saved_files"][0]
            self.assertEqual(saved["assignment"], "capacity_3")
            self.assertTrue(Path(saved["saved_path"]).exists())
            self.assertTrue(Path(saved["processed_saved_path"]).exists())
            self.assertIn("capacity_3_cyc", saved["saved_path"])
            self.assertEqual(len(commit_payload["match_overrides"]), 1)
            override_row = commit_payload["match_overrides"][0]
            self.assertEqual(override_row["kind"], "capacity")
            self.assertEqual(override_row["journal_row"], 2)
            override_data = json.loads(capacity_override_path.read_text(encoding="utf-8"))
            override_key = str(Path(saved["processed_saved_path"]).resolve().relative_to((root / "capacity").resolve()))
            self.assertEqual(override_row["relative_path"], override_key)
            self.assertEqual(override_data[override_key]["condition_key"], "sample 1")
            self.assertEqual(override_data[override_key]["journal_row"], 2)
            self.assertEqual(override_data[override_key]["selection_source"], "import_commit")
            self.assertEqual(override_data[override_key]["import_draft_id"], payload["draft_id"])
            persist_by_kind = {row["kind"]: row for row in commit_payload["persist_outputs"]}
            self.assertTrue(persist_by_kind["summary_metrics"]["ok"])
            self.assertTrue(Path(persist_by_kind["summary_metrics"]["path"]).exists())
            self.assertTrue(persist_by_kind["capacity_match_outputs"]["ok"])
            self.assertTrue((root / "battery_visual_outputs" / "capacity_match_report.json").exists())
            self.assertTrue((root / "battery_visual_outputs" / "capacity_condition_matches.csv").exists())
            self.assertEqual(match_refresh_response.status_code, 200)
            match_refresh_payload = match_refresh_response.get_json()
            self.assertTrue(match_refresh_payload["report_ready"])
            refreshed = {
                row["relative_path"]: row
                for row in match_refresh_payload["final_rows"]
                if row.get("relative_path") == override_key
            }
            self.assertEqual(refreshed[override_key]["status"], "manual")
            self.assertEqual(refreshed[override_key]["journal_row"], 2)

            reloaded = load_workbook(workbook_path, data_only=False)
            try:
                sheet = reloaded["JYJ"]
                self.assertEqual(sheet.cell(row=2, column=1).value, "sample 1")
                self.assertEqual(sheet.cell(row=2, column=2).value, "260627")
                self.assertEqual(sheet.cell(row=2, column=3).value, "LIB")
            finally:
                reloaded.close()

    def test_append_and_delete_draft_files_via_api(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            app = Flask(__name__)
            app.secret_key = "test"
            register_battery_lab(app)
            client = app.test_client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"):
                created = client.post(
                    "/battery/api/import/drafts",
                    data={"files": (io.BytesIO(synthetic_wrd_bytes()), "first.wrd")},
                    content_type="multipart/form-data",
                ).get_json()
                draft_id = created["draft_id"]
                self.assertEqual(len(created["files"]), 1)
                # All five types are offered in the toggle; exclude is removed (delete button replaces it).
                self.assertEqual(
                    created["files"][0]["assignment_options"],
                    ["eis_comparison", "eis_time_series", "capacity_1", "capacity_2", "capacity_3"],
                )

                appended = client.post(
                    f"/battery/api/import/drafts/{draft_id}/files",
                    data={"files": (io.BytesIO(synthetic_wrd_bytes()), "second.wrd")},
                    content_type="multipart/form-data",
                ).get_json()
                self.assertTrue(appended["ok"])
                self.assertEqual(len(appended["files"]), 2)
                first_id = appended["files"][0]["file_id"]
                first_raw = Path(appended["files"][0]["raw_path"])
                self.assertTrue(first_raw.exists())

                deleted = client.delete(f"/battery/api/import/drafts/{draft_id}/files/{first_id}")
                self.assertEqual(deleted.status_code, 200)
                deleted_payload = deleted.get_json()
                self.assertTrue(deleted_payload["ok"])
                self.assertEqual(len(deleted_payload["files"]), 1)
                self.assertNotEqual(deleted_payload["files"][0]["file_id"], first_id)
                self.assertFalse(first_raw.exists())

                missing = client.delete(f"/battery/api/import/drafts/{draft_id}/files/does-not-exist")
                self.assertEqual(missing.status_code, 400)

    def test_draft_overlay_uses_live_viewer_pipeline(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            app = Flask(__name__)
            app.secret_key = "test"
            register_battery_lab(app)
            client = app.test_client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"):
                created = client.post(
                    "/battery/api/import/drafts",
                    data={"files": (io.BytesIO(synthetic_wrd_bytes()), "260627_A_0.1C.wrd")},
                    content_type="multipart/form-data",
                ).get_json()
                draft_id = created["draft_id"]
                file_id = created["files"][0]["file_id"]
                overlay = client.get(
                    f"/battery/api/import/drafts/{draft_id}/overlay"
                    f"?file_ids={file_id}&kind=capacity&color_mode=comparison&title=Capacity%201"
                )
                self.assertEqual(overlay.status_code, 200)
                payload = overlay.get_json()
                self.assertTrue(payload["available"])
                self.assertGreaterEqual(payload["series_count"], 1)
                self.assertIn("<svg", payload["html"])


if __name__ == "__main__":
    unittest.main()
