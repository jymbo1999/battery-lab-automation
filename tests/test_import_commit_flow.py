"""Per-unit import flow: each file/unit -> its own journal row.

Covers the redesigned '새 실험 등록' wizard backend (field-spec, per-unit
metadata, normalized-name preview, one journal row per row-unit, literal
derived densities, protocol-token capacity names, graph artifacts on commit).
"""
import io
import tempfile
import unittest
from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

from flask import Flask
from openpyxl import Workbook, load_workbook

from battery_lab import register_battery_lab, routes
from test_experiment_import import synthetic_wrd_bytes


JYJ_HEADERS = [
    "참고", "전해질", "종류", "Date", "Sample", "Conductive agent", "Binder", "Current (A)",
    "Active material (g)", "CV (uA)", "Cut capacity (Ah)", "Voltage range", "Cell 자리",
    "Theoretical capacity (mAh/g)", "C-Rate (1/h)", "foil+electrode (g)", "foil (g)", "ratio",
    "Current density (mA/g)", "Areal mass density (mg/cｍ2)", "전극(foil+electrode) 두께(mm)",
    "호일 두께(mm)", "전극 두께(mm)", "electrode(g)", "volume (mm3)", "합제밀도(g/cm3)",
]

# Full field set the redesigned form sends: 4 user-filled + 12 fixed defaults.
def full_meta(sample: str, *, foil_electrode_g: str = "0.0150") -> dict:
    return {
        "date": "260627", "sample": sample, "foil_electrode_g": foil_electrode_g,
        "foil_electrode_mm": "0.020", "reference": "12 파이_Cu foil",
        "electrolyte": "1.0M LiPF6 EC/DEC 1:1", "cell_type": "LIB", "conductive_agent": "-",
        "binder": "2wt%cmc", "voltage_range": "0.01~2V", "foil_g": "0.009928", "ratio": "0.96",
        "current_density": "37.2", "foil_thickness_mm": "0.00958", "electrolyte_ul": "80",
        "drying_condition": "60도 12시간",
    }


def make_journal(root: Path) -> Path:
    path = root / "Project_Abstract" / "Cell condition Calculation.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "JYJ"
    worksheet.append(JYJ_HEADERS)
    workbook.save(path)
    workbook.close()
    return path


@contextmanager
def committed_roots(root: Path, workbook_path: Path):
    with (
        patch.object(routes, "BATTERY_CONDITION_WORKBOOK", workbook_path),
        patch.object(routes, "BATTERY_CAPACITY_ROOT", root / "capacity"),
        patch.object(routes, "BATTERY_EIS_ROOT", root / "EIS"),
        patch.object(routes, "BATTERY_MATCH_CAPACITY_JSON", root / "battery_visual_outputs" / "capacity_match_overrides.json"),
        patch.object(routes, "BATTERY_MATCH_EIS_JSON", root / "battery_visual_outputs" / "eis_match_overrides.json"),
    ):
        yield


class PerUnitImportFlowTests(unittest.TestCase):
    def _client(self):
        app = Flask(__name__)
        app.secret_key = "test"
        register_battery_lab(app)
        return app.test_client()

    def _upload(self, client, filename):
        return client.post(
            "/battery/api/import/drafts",
            data={"files": (io.BytesIO(synthetic_wrd_bytes()), filename)},
            content_type="multipart/form-data",
        ).get_json()

    def test_field_spec_exposes_16_fields_and_binder_presets(self):
        client = self._client()
        with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", Path(tempfile.mkdtemp()) / "out"):
            payload = client.get("/battery/api/import/field-spec").get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(len(payload["fields"]), 16)
        self.assertEqual(payload["binder_presets"], ["2wt%cmc", "2wt%cmc/40wt%SBR"])

    def test_single_file_commits_one_row_with_literal_density_and_protocol_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            client = self._client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"):
                created = self._upload(client, "sample.wrd")
                draft_id = created["draft_id"]
                file_id = created["files"][0]["file_id"]
                # one capacity file => one row-unit keyed by its file_id
                self.assertEqual([u["unit_id"] for u in created["units"]], [file_id])

                workbook_path = make_journal(root)
                with committed_roots(root, workbook_path):
                    meta = client.patch(
                        f"/battery/api/import/drafts/{draft_id}/units/{file_id}/metadata",
                        json={"metadata": full_meta("cell A")},
                    ).get_json()
                    self.assertTrue(meta["ok"])
                    self.assertEqual(meta["unit_metadata_status"], "ready")

                    names = client.get(f"/battery/api/import/drafts/{draft_id}/normalized-names").get_json()
                    self.assertEqual(names["rows"][0]["raw_name"], "sample.wrd")
                    # predicted row 2, capacity_1 -> 0.1C protocol token
                    self.assertEqual(names["rows"][0]["normalized_name"], "2_cell_A_0.1C.wrd")

                    commit = client.post(f"/battery/api/import/drafts/{draft_id}/commit").get_json()

            self.assertTrue(commit["ok"])
            self.assertEqual(commit["journal_rows"], [2])
            self.assertEqual(len(commit["saved_files"]), 1)
            saved = commit["saved_files"][0]
            self.assertEqual(saved["journal_row"], 2)
            self.assertIn("0.1C", Path(saved["saved_path"]).name)  # protocol token, not capacity_1
            self.assertTrue(Path(saved["saved_path"]).exists())

            # Problem A: areal mass density readable as a numeric literal pre-Excel-recalc.
            workbook = load_workbook(workbook_path, data_only=True)
            try:
                sheet = workbook["JYJ"]
                self.assertEqual(sheet.cell(row=2, column=5).value, "cell A")  # Sample
                self.assertIsInstance(sheet.cell(row=2, column=20).value, float)  # Areal mass density
                self.assertGreater(sheet.cell(row=2, column=20).value, 0)
                self.assertEqual(sheet.cell(row=2, column=22).value, "0.00958")  # 호일 두께 (no 'mm' collision)
            finally:
                workbook.close()

            # match override pinned to the new row
            self.assertEqual([(o["kind"], o["journal_row"]) for o in commit["match_overrides"]], [("capacity", 2)])

            # Task 13: graph artifacts generated on commit
            plots = [row for row in commit["persist_outputs"] if row["kind"] == "plot" and row["ok"]]
            self.assertTrue(plots, "commit should regenerate at least one plot artifact")
            self.assertTrue(Path(plots[0]["artifact_path"]).exists())
            self.assertTrue((root / "battery_visual_outputs" / "capacity_match_report.json").exists())

    def test_two_files_create_two_independent_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            client = self._client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"):
                created = self._upload(client, "first.wrd")
                draft_id = created["draft_id"]
                appended = client.post(
                    f"/battery/api/import/drafts/{draft_id}/files",
                    data={"files": (io.BytesIO(synthetic_wrd_bytes()), "second.wrd")},
                    content_type="multipart/form-data",
                ).get_json()
                units = appended["units"]
                self.assertEqual(len(units), 2)  # strict per-file: two capacity files => two units

                workbook_path = make_journal(root)
                with committed_roots(root, workbook_path):
                    for index, unit in enumerate(units):
                        resp = client.patch(
                            f"/battery/api/import/drafts/{draft_id}/units/{unit['unit_id']}/metadata",
                            json={"metadata": full_meta(f"cell {index}")},
                        ).get_json()
                        self.assertTrue(resp["ok"], resp.get("unit_metadata_errors"))
                    commit = client.post(f"/battery/api/import/drafts/{draft_id}/commit").get_json()

            self.assertTrue(commit["ok"])
            self.assertEqual(sorted(commit["journal_rows"]), [2, 3])  # one row per unit
            self.assertEqual(len(commit["saved_files"]), 2)
            self.assertEqual({s["journal_row"] for s in commit["saved_files"]}, {2, 3})

            workbook = load_workbook(workbook_path, data_only=True)
            try:
                sheet = workbook["JYJ"]
                self.assertEqual(sheet.max_row, 3)  # header + 2 rows
                self.assertEqual({sheet.cell(row=r, column=5).value for r in (2, 3)}, {"cell 0", "cell 1"})
            finally:
                workbook.close()

    def test_commit_blocked_until_all_units_ready(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            client = self._client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", root / "battery_visual_outputs"):
                created = self._upload(client, "only.wrd")
                draft_id = created["draft_id"]
                workbook_path = make_journal(root)
                with committed_roots(root, workbook_path):
                    # no metadata set for the unit -> commit must fail
                    commit = client.post(f"/battery/api/import/drafts/{draft_id}/commit")
            self.assertEqual(commit.status_code, 400)
            self.assertIn("ready", commit.get_json()["error"])


if __name__ == "__main__":
    unittest.main()
