"""Row-detail file-replace flow: backup + drop-in + full recompute, pinned to row.

Reuses the per-unit import fixtures to register one capacity file to a journal
row, then exercises POST /api/journal/row-replace-file and asserts the old file
is backed up + retired, the new file is placed, the match override is repointed
to the new file (same journal row), and the experiment-info cells are untouched.
"""
import io
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from battery_lab import routes
from test_experiment_import import synthetic_wrd_bytes
from test_import_commit_flow import (
    PerUnitImportFlowTests,
    committed_roots,
    full_meta,
    make_journal,
)


class RowReplaceFlowTests(unittest.TestCase):
    def _client(self):
        return PerUnitImportFlowTests()._client()

    def _commit_one(self, client, root, workbook_path, filename="orig.wrd"):
        created = client.post(
            "/battery/api/import/drafts",
            data={"files": (io.BytesIO(synthetic_wrd_bytes()), filename)},
            content_type="multipart/form-data",
        ).get_json()
        draft_id = created["draft_id"]
        file_id = created["files"][0]["file_id"]
        client.patch(
            f"/battery/api/import/drafts/{draft_id}/units/{file_id}/metadata",
            json={"metadata": full_meta("cell A")},
        )
        return client.post(f"/battery/api/import/drafts/{draft_id}/commit").get_json()

    def test_replace_backs_up_old_and_repoints_override(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            output_root = root / "battery_visual_outputs"
            capacity_override = output_root / "capacity_match_overrides.json"
            client = self._client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", output_root):
                workbook_path = make_journal(root)
                with committed_roots(root, workbook_path):
                    commit = self._commit_one(client, root, workbook_path)
                    self.assertTrue(commit["ok"])
                    self.assertEqual(commit["journal_rows"], [2])

                    overrides = json.loads(capacity_override.read_text())
                    self.assertEqual(len(overrides), 1)
                    old_rel = next(iter(overrides))
                    old_abs = root / "capacity" / old_rel
                    self.assertTrue(old_abs.exists())

                    resp = client.post(
                        "/battery/api/journal/row-replace-file",
                        data={
                            "row": "2",
                            "kind": "capacity",
                            "target": old_rel,
                            "file": (io.BytesIO(synthetic_wrd_bytes()), "fixed.wrd"),
                        },
                        content_type="multipart/form-data",
                    )
                    result = resp.get_json()

            self.assertEqual(resp.status_code, 200, result)
            self.assertTrue(result["ok"], result)
            self.assertEqual(result["row"], 2)
            self.assertEqual(result["kind"], "capacity")

            # New matched file exists; old one retired.
            new_abs = root / "capacity" / result["new_rel_path"]
            self.assertTrue(new_abs.exists())
            self.assertFalse(old_abs.exists(), "old matched file should be retired")

            # Old file is recoverable from the backup dir.
            backup_dir = Path(result["backup_dir"])
            self.assertTrue(backup_dir.exists())
            self.assertTrue(any(backup_dir.iterdir()), "backup dir should contain the retired file(s)")

            # Override repointed to the new file, still pinned to row 2.
            overrides = json.loads(capacity_override.read_text())
            self.assertNotIn(old_rel, overrides)
            self.assertIn(result["new_rel_path"], overrides)
            self.assertEqual(overrides[result["new_rel_path"]]["journal_row"], 2)
            self.assertEqual(overrides[result["new_rel_path"]]["selection_source"], "row_file_replace")

            # Recompute ran and produced a fresh match report.
            self.assertTrue(any(r.get("ok") for r in result["recompute"]))
            self.assertTrue((output_root / "capacity_match_report.json").exists())

            # Experiment-info cells untouched (Sample stays 'cell A').
            workbook = load_workbook(workbook_path, data_only=True)
            try:
                self.assertEqual(workbook["JYJ"].cell(row=2, column=5).value, "cell A")
            finally:
                workbook.close()

    def test_replace_rejects_unknown_target(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "battery"
            output_root = root / "battery_visual_outputs"
            client = self._client()
            with client, patch.object(routes, "BATTERY_OUTPUT_ROOT", output_root):
                workbook_path = make_journal(root)
                with committed_roots(root, workbook_path):
                    self._commit_one(client, root, workbook_path)
                    resp = client.post(
                        "/battery/api/journal/row-replace-file",
                        data={
                            "row": "2",
                            "kind": "capacity",
                            "target": "does/not/exist.csv",
                            "file": (io.BytesIO(synthetic_wrd_bytes()), "fixed.wrd"),
                        },
                        content_type="multipart/form-data",
                    )
            self.assertEqual(resp.status_code, 400)
            self.assertFalse(resp.get_json()["ok"])


if __name__ == "__main__":
    unittest.main()
