import unittest
from battery_lab.experiment_import import (
    IMPORT_JOURNAL_FIELDS, field_keys, fixed_defaults, variable_keys,
)


class ImportJournalFieldsTests(unittest.TestCase):
    def test_field_spec_has_16_fields_4_variable_12_fixed(self):
        self.assertEqual(len(IMPORT_JOURNAL_FIELDS), 16)
        self.assertEqual(len(variable_keys()), 4)
        self.assertEqual(sum(1 for f in IMPORT_JOURNAL_FIELDS if f["bucket"] == "fixed"), 12)

    def test_variable_fields_are_blank_user_filled(self):
        self.assertEqual(
            variable_keys(),
            ["date", "sample", "foil_electrode_g", "foil_electrode_mm"],
        )

    def test_fixed_defaults_match_spec(self):
        d = fixed_defaults()
        self.assertEqual(d["reference"], "12 파이_Cu foil")
        self.assertEqual(d["electrolyte"], "1.0M LiPF6 EC/DEC 1:1")
        self.assertEqual(d["cell_type"], "LIB")
        self.assertEqual(d["foil_g"], "0.009928")
        self.assertEqual(d["ratio"], "0.96")
        self.assertEqual(d["current_density"], "37.2")
        self.assertEqual(d["foil_thickness_mm"], "0.00958")
        self.assertEqual(d["drying_condition"], "60도 12시간")

    def test_each_field_maps_to_exact_excel_header(self):
        headers = {f["key"]: f["header"] for f in IMPORT_JOURNAL_FIELDS}
        self.assertEqual(headers["foil_thickness_mm"], "호일 두께(mm)")
        self.assertEqual(headers["foil_electrode_mm"], "전극(foil+electrode) 두께(mm)")
        self.assertEqual(headers["reference"], "참고")
        self.assertEqual(headers["cell_type"], "종류")


import math
from battery_lab.experiment_import import compute_derived_metadata


class ComputeDerivedTests(unittest.TestCase):
    def test_areal_mass_density_from_foil_inputs(self):
        derived = compute_derived_metadata(
            {"foil_electrode_g": "0.0150", "foil_g": "0.009928", "ratio": "0.96",
             "foil_electrode_mm": "0.020", "foil_thickness_mm": "0.00958"}
        )
        active = (0.0150 - 0.009928) * 0.96
        self.assertAlmostEqual(derived["active_material_g"], active, places=9)
        self.assertAlmostEqual(
            derived["areal_mass_density"], active * 1000 / (math.pi * 0.6 ** 2), places=6
        )

    def test_electrode_density_from_thickness(self):
        derived = compute_derived_metadata(
            {"foil_electrode_g": "0.0150", "foil_g": "0.009928", "ratio": "0.96",
             "foil_electrode_mm": "0.020", "foil_thickness_mm": "0.00958"}
        )
        electrode_g = 0.0150 - 0.009928
        thickness = 0.020 - 0.00958
        volume = 113.1 * thickness
        self.assertAlmostEqual(derived["electrode_density"], electrode_g / (volume / 1000), places=6)

    def test_missing_inputs_yield_none(self):
        self.assertIsNone(compute_derived_metadata({"foil_g": "0.009928"})["areal_mass_density"])


from battery_lab.experiment_import import validate_metadata, clean_metadata


class ValidateMetadataTests(unittest.TestCase):
    def _valid(self):
        return {
            "date": "260627", "sample": "cell A",
            "foil_electrode_g": "0.0150", "foil_electrode_mm": "0.020",
            "foil_g": "0.009928", "ratio": "0.96",
        }

    def test_valid_metadata_has_no_errors(self):
        self.assertEqual(validate_metadata(self._valid()), [])

    def test_required_variable_fields_enforced(self):
        m = self._valid(); del m["foil_electrode_g"]
        self.assertIn("foil_electrode_g is required", validate_metadata(m))

    def test_foil_electrode_must_exceed_foil(self):
        m = self._valid(); m["foil_electrode_g"] = "0.005"
        self.assertTrue(any("foil+electrode" in e for e in validate_metadata(m)))

    def test_ratio_range(self):
        m = self._valid(); m["ratio"] = "1.5"
        self.assertTrue(any("ratio" in e for e in validate_metadata(m)))

    def test_clean_metadata_keeps_only_spec_keys(self):
        cleaned = clean_metadata({**self._valid(), "sample_group": "x", "junk": "y"})
        self.assertNotIn("sample_group", cleaned)
        self.assertNotIn("junk", cleaned)
        self.assertIn("sample", cleaned)


import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from battery_lab.experiment_import import append_journal_row


class JournalWriterTests(unittest.TestCase):
    def _make_book(self, path):
        wb = Workbook(); ws = wb.active; ws.title = "JYJ"
        ws.append(["참고", "전해질", "종류", "Date", "Sample", "Conductive agent", "Binder",
                   "Current (A)", "Active material (g)", "CV (uA)", "Cut capacity (Ah)", "Voltage range",
                   "Cell 자리", "Theoretical capacity (mAh/g)", "C-Rate (1/h)", "foil+electrode (g)",
                   "foil (g)", "ratio", "Current density (mA/g)", "Areal mass density (mg/cｍ2)",
                   "전극(foil+electrode) 두께(mm)", "호일 두께(mm)", "전극 두께(mm)", "electrode(g)",
                   "volume (mm3)", "합제밀도(g/cm3)"])
        wb.save(path); wb.close()

    def test_writes_thickness_to_correct_distinct_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "j.xlsx"; self._make_book(p)
            metadata = {"sample": "A", "foil_electrode_g": "0.0150", "foil_g": "0.009928",
                        "ratio": "0.96", "foil_electrode_mm": "0.020", "foil_thickness_mm": "0.00958",
                        "reference": "12 파이_Cu foil"}
            row = append_journal_row(p, "JYJ", metadata)
            wb = load_workbook(p); ws = wb["JYJ"]
            self.assertEqual(row, 2)
            self.assertEqual(ws.cell(row=2, column=21).value, "0.020")  # 전극(foil+electrode) 두께(mm)
            self.assertEqual(ws.cell(row=2, column=22).value, "0.00958")  # 호일 두께(mm) — no 'mm' collision
            self.assertEqual(ws.cell(row=2, column=1).value, "12 파이_Cu foil")
            wb.close()

    def test_areal_density_written_as_numeric_literal(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "j.xlsx"; self._make_book(p)
            metadata = {"sample": "A", "foil_electrode_g": "0.0150", "foil_g": "0.009928", "ratio": "0.96"}
            append_journal_row(p, "JYJ", metadata)
            wb = load_workbook(p, data_only=True); ws = wb["JYJ"]
            self.assertIsInstance(ws.cell(row=2, column=20).value, float)  # Areal mass density literal
            self.assertGreater(ws.cell(row=2, column=20).value, 0)
            wb.close()


from battery_lab.experiment_import import assignment_protocol_token


class ProtocolNamingTests(unittest.TestCase):
    def test_assignment_maps_to_human_protocol_token(self):
        self.assertEqual(assignment_protocol_token("capacity_1"), "0.1C")
        self.assertEqual(assignment_protocol_token("capacity_2"), "0.5C")
        self.assertEqual(assignment_protocol_token("capacity_3"), "rate per")
        self.assertEqual(assignment_protocol_token("eis_comparison"), "eis_comparison")


if __name__ == "__main__":
    unittest.main()
