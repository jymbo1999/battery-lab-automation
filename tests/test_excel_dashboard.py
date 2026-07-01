import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from battery_lab.excel_dashboard import WorkbookStore, build_sheet_payload, render_page


def make_formula_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "JYJ"
    headers = {
        8: "Current (A)",
        9: "Active material (g)",
        16: "전극+Cu foil (g)",
        17: "Cu foil (g)",
        18: "활물질 비율",
        19: "용량",
        20: "Areal mass density (mg/cｍ2)",
        21: "전극(foil+electrode) 두께(mm)",
        22: "foil 두께(mm)",
        23: "전극 두께(mm)",
        24: "electrode(g)",
        25: "volume (mm3)",
        26: "합제밀도(g/cm3)",
    }
    for column, header in headers.items():
        worksheet.cell(row=1, column=column, value=header)
    worksheet["P2"] = 0.020
    worksheet["Q2"] = 0.010
    worksheet["R2"] = 0.8
    worksheet["S2"] = 186
    worksheet["U2"] = 0.08
    worksheet["V2"] = 0.01
    return workbook


class ExcelDashboardTests(unittest.TestCase):
    def test_sheet_payload_preserves_values_dimensions_and_styles(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JYJ"
        worksheet["A1"] = "참고"
        worksheet["A1"].font = Font(bold=True, color="FFFFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor="FF4472C4")
        worksheet["B2"] = "merged"
        worksheet.merge_cells("B2:C3")
        worksheet.column_dimensions["A"].width = 15
        worksheet.row_dimensions[1].height = 27

        payload = build_sheet_payload(worksheet, Path("conditions.xlsx"))

        self.assertEqual(payload["sheet"], "JYJ")
        self.assertEqual(payload["sourceMaxRow"], 3)
        self.assertEqual(payload["extraRows"], 100)
        self.assertEqual(payload["extraStartRow"], 4)
        self.assertEqual(payload["maxRow"], 103)
        self.assertEqual(payload["rows"][-1]["index"], 103)
        self.assertTrue(payload["rows"][-1]["extra"])
        self.assertFalse(payload["rows"][-1]["ignored"])
        self.assertGreater(payload["columns"][0]["width"], 100)
        self.assertEqual(payload["rows"][0]["height"], 36)
        self.assertEqual(payload["rows"][0]["cells"][0]["value"], "참고")
        self.assertTrue(payload["rows"][0]["cells"][0]["style"]["bold"])
        self.assertEqual(payload["rows"][0]["cells"][0]["style"]["backgroundColor"], "#4472C4")
        merged_cell = next(cell for cell in payload["rows"][1]["cells"] if cell["address"] == "B2")
        self.assertEqual((merged_cell["rowspan"], merged_cell["colspan"]), (2, 2))

    def test_sheet_payload_marks_rows_outside_condition_filter(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JYJ"
        headers = ["참고", "전해질", "종류", "Binder", "Voltage range"]
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column, value=header)
        valid = ["12 파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"]
        invalid = ["12 파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "5wt% pvdf/nmp", "0.01~2V"]
        for column, value in enumerate(valid, start=1):
            worksheet.cell(row=2, column=column, value=value)
        for column, value in enumerate(invalid, start=1):
            worksheet.cell(row=3, column=column, value=value)

        payload = build_sheet_payload(worksheet, Path("conditions.xlsx"))

        self.assertFalse(payload["rows"][1]["ignored"])
        self.assertTrue(payload["rows"][2]["ignored"])
        self.assertEqual(payload["filter"]["matchedRows"], 1)
        self.assertEqual(payload["filter"]["ignoredRows"], 1)

    def test_sheet_payload_can_limit_visible_rows_for_fast_initial_view(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JYJ"
        worksheet.append(["Sample", "참고", "전해질", "종류", "Binder", "Voltage range"])
        for row_idx in range(1, 8):
            worksheet.append([f"matched-{row_idx}", "12파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"])

        payload = build_sheet_payload(worksheet, Path("conditions.xlsx"), include_ignored=False, row_limit=3, extra_rows=2)

        self.assertTrue(payload["partialRows"])
        self.assertEqual(payload["sourceMaxRow"], 8)
        self.assertEqual(payload["extraRows"], 2)
        self.assertEqual(payload["extraStartRow"], 9)
        self.assertEqual([row["index"] for row in payload["rows"]], [1, 6, 7, 8, 9, 10])

    def test_filtered_extra_rows_start_after_last_populated_row_not_hidden_data(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JYJ"
        worksheet.append(["Sample", "참고", "전해질", "종류", "Binder", "Voltage range"])
        worksheet.append(["matched-1", "12파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"])
        worksheet.append(["matched-2", "12파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"])
        worksheet.append(["ignored-1", "12파이_Ti foil", "2M ZnSO4", "ZIB", "PVDF", "0.01~2V"])
        worksheet.append(["ignored-2", "12파이_Ti foil", "2M ZnSO4", "ZIB", "PVDF", "0.01~2V"])
        worksheet.row_dimensions[3].height = 25
        worksheet["A3"].fill = PatternFill("solid", fgColor="FFD9EAD3")

        payload = build_sheet_payload(worksheet, Path("conditions.xlsx"), include_ignored=False, extra_rows=2)

        self.assertEqual(payload["sourceMaxRow"], 5)
        # Ignored ZIB rows 4-5 hold real data but are filtered out; blank editable
        # rows must begin AFTER the last populated row (5) so they never overdraw
        # the hidden data. Rows 4-5 are not rendered in hide mode.
        self.assertEqual(payload["extraStartRow"], 6)
        self.assertEqual(payload["maxRow"], 7)
        self.assertEqual([row["index"] for row in payload["rows"]], [1, 2, 3, 6, 7])
        self.assertFalse(payload["rows"][2]["extra"])
        self.assertTrue(payload["rows"][3]["extra"])
        self.assertEqual(payload["rows"][3]["height"], 33)
        self.assertEqual(payload["rows"][3]["cells"][0]["address"], "A6")
        self.assertEqual(payload["rows"][3]["cells"][0]["value"], "")
        self.assertEqual(payload["rows"][3]["cells"][0]["style"]["backgroundColor"], "#D9EAD3")

    def test_filtered_extra_rows_recalibrate_after_new_matching_row_is_added(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JYJ"
        worksheet.append(["Sample", "참고", "전해질", "종류", "Binder", "Voltage range"])
        worksheet.append(["matched-1", "12파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"])
        worksheet.append(["ignored", "12파이_Ti foil", "2M ZnSO4", "ZIB", "PVDF", "0.01~2V"])
        before = build_sheet_payload(worksheet, Path("conditions.xlsx"), include_ignored=False, extra_rows=2)

        worksheet.append(["matched-2", "12파이_Cu foil", "1.0M LiPF6 EC/DEC 1:1", "LIB", "2wt% cmc", "0.01~2V"])
        after = build_sheet_payload(worksheet, Path("conditions.xlsx"), include_ignored=False, extra_rows=2)

        # Before: the trailing ignored ZIB row (3) holds data, so blank editable
        # rows start after it (4), and row 3 stays hidden rather than overdrawn.
        self.assertEqual(before["extraStartRow"], 4)
        self.assertEqual([row["index"] for row in before["rows"] if row["extra"]], [4, 5])
        self.assertNotIn(3, [row["index"] for row in before["rows"]])
        # After appending a matched row (4), extra rows recalibrate below it.
        self.assertEqual(after["extraStartRow"], 5)
        self.assertEqual([row["index"] for row in after["rows"] if row["extra"]], [5, 6])
        self.assertFalse([row for row in after["rows"] if row["index"] == 4][0]["extra"])

    def test_store_updates_workbook_cell_value(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "conditions.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "JYJ"
            worksheet["A1"] = "Sample"
            workbook.save(path)

            store = WorkbookStore(path, "JYJ")
            result = store.update_cell(1, 1, "updated")

            self.assertEqual(result["value"], "updated")
            saved = load_workbook(path)
            self.assertEqual(saved["JYJ"]["A1"].value, "updated")
            saved.close()

    def test_sheet_payload_computes_excel_style_formula_columns(self):
        workbook = make_formula_workbook()
        worksheet = workbook.active

        payload = build_sheet_payload(worksheet, Path("conditions.xlsx"))

        row = payload["rows"][1]
        cells = {cell["address"]: cell for cell in row["cells"]}
        self.assertEqual(cells["I2"]["formula"], "=(P2-Q2)*R2")
        self.assertTrue(cells["I2"]["formulaCell"])
        self.assertFalse(cells["I2"]["editable"])
        self.assertAlmostEqual(float(cells["I2"]["value"]), 0.008)
        self.assertAlmostEqual(float(cells["H2"]["value"]), 0.001488)
        self.assertAlmostEqual(float(cells["T2"]["value"]), 0.008 * 1000 / (3.141592653589793 * 0.6 * 0.6))
        self.assertAlmostEqual(float(cells["W2"]["value"]), 0.07)
        self.assertAlmostEqual(float(cells["X2"]["value"]), 0.01)
        self.assertAlmostEqual(float(cells["Y2"]["value"]), 113.1 * 0.07)
        self.assertAlmostEqual(float(cells["Z2"]["value"]), 0.01 / ((113.1 * 0.07) / 1000))

    def test_store_rewrites_dependent_formula_columns_after_edit(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "conditions.xlsx"
            workbook = make_formula_workbook()
            workbook.save(path)

            store = WorkbookStore(path, "JYJ")
            store.update_cell(2, 16, "0.03")

            saved = load_workbook(path, data_only=False)
            sheet = saved["JYJ"]
            self.assertEqual(sheet["H2"].value, "=I2*S2*1000/10^6")
            self.assertEqual(sheet["I2"].value, "=(P2-Q2)*R2")
            self.assertEqual(sheet["T2"].value, "=I2*1000/(PI()*(0.6)^2)")
            self.assertEqual(sheet["W2"].value, "=U2-V2")
            self.assertEqual(sheet["X2"].value, "=P2-Q2")
            self.assertEqual(sheet["Y2"].value, "=113.1*W2")
            self.assertEqual(sheet["Z2"].value, "=X2/(Y2/1000)")
            saved.close()

    def test_render_page_contains_selectable_copy_grid_hooks(self):
        page = render_page()

        self.assertIn("/api/sheet", page)
        self.assertIn("/api/cell", page)
        self.assertIn("contentEditable = 'false'", page)
        self.assertNotIn("contentEditable = cell.editable ? 'true' : 'false'", page)
        self.assertIn("selected-cell", page)
        self.assertIn("selection-anchor", page)
        self.assertIn("copySelectedCells", page)
        self.assertIn("selectedCellMatrix", page)
        self.assertIn("event.clipboardData.setData('text/plain', text)", page)
        self.assertIn(r"const text = rows.map(row => row.join('\t')).join('\n')", page)
        self.assertNotIn("row.join('\t')).join('\n", page)
        self.assertIn("loadingFilteredRows", page)
        self.assertIn("loadCompleteFilteredRows", page)
        self.assertIn("if (data.partialRows && options.fastView) loadCompleteFilteredRows()", page)
        self.assertIn("fetch(sheetUrl(false, false))", page)
        self.assertIn("if (state.fullDataLoaded || state.filterMode !== 'hide')", page)
        self.assertIn("Loading all matched rows", page)
        self.assertIn("selectCells(state.selection.anchorRow, state.selection.anchorColumn, cell.row, cell.column)", page)
        self.assertIn("filterMode", page)
        self.assertIn("실험 일지", page)
        self.assertIn("무시 행 표시안함", page)
        self.assertIn("무시 행 회색", page)
        self.assertIn("필터 기준", page)
        self.assertIn("header-row", page)
        self.assertIn("zoomOut", page)
        self.assertIn("zoomIn", page)
        self.assertIn("zoom: 0.55", page)
        self.assertIn("<output id=\"zoomValue\">55%</output>", page)
        self.assertIn("scrollToLatestRows", page)
        self.assertIn("extraStartRow", page)
        self.assertIn("visibleExtraRows = 6", page)
        self.assertIn("tr.dataset.rowIndex = row.index", page)
        self.assertIn("state.zoom", page)
        self.assertIn("viewportCenterAnchor", page)
        self.assertIn("contentX: (sheetEl.scrollLeft + viewportX) / oldZoom", page)
        self.assertIn("setZoom(state.zoom + step, { clientX: event.clientX, clientY: event.clientY })", page)
        self.assertIn("formula-cell", page)
        self.assertIn("dataset.formula", page)


if __name__ == "__main__":
    unittest.main()
