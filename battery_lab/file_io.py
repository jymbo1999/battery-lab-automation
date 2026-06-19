from __future__ import annotations

import csv
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from .models import FileMeta, ParsedDataset

try:
    from eis_fit_handoff.eis_circle_fit import fit_eis_first_arc, load_valid_fit_metadata, save_fit_metadata
except ModuleNotFoundError:  # pragma: no cover - keeps non-EIS use working in minimal installs.
    fit_eis_first_arc = None
    load_valid_fit_metadata = None
    save_fit_metadata = None

try:
    from wonatech_parsers.eis import parse_eis_file as parse_wonatech_eis_file
except ModuleNotFoundError:  # pragma: no cover
    parse_wonatech_eis_file = None


ANALYSIS_CAPACITY = "capacity"
ANALYSIS_VOLTAGE = "voltage_profile"
ANALYSIS_EIS = "eis"
ANALYSIS_SHEET = "sheet_resistance"
ANALYSIS_RAMAN = "raman"
ANALYSIS_TGA = "tga"
ANALYSIS_UNKNOWN = "unknown"

ALIASES = {
    "cycle": {"cycle", "cycle number", "cycle no", "cyc#", "cyc"},
    "charge_capacity": {
        "charge capacity",
        "charge_capacity",
        "chg cap",
        "charge cap",
        "charge/mah/g",
        "q_ch/m [mah/g]",
        "q_charge [mah]",
        "q_charge_mah",
        "q_charge_mah_g",
    },
    "discharge_capacity": {
        "discharge capacity",
        "discharge_capacity",
        "dchg cap",
        "discharge cap",
        "discharge/mah/g",
        "q_dis/m [mah/g]",
        "q_discharge [mah]",
        "q_discharge_mah",
        "q_discharge_mah_g",
    },
    "capacity": {"capacity", "cap", "capacity/mah/g", "specific capacity", "q/m [mah/g]", "q [mah]"},
    "voltage": {"voltage", "voltage_v", "ewe/v", "ewe", "v", "potential", "v [v]"},
    "direction": {"direction", "type", "step type", "mode"},
    "frequency": {"frequency", "frequency_hz", "freq", "freq/hz", "hz"},
    "z_real": {"zreal", "zreal_ohm", "z real", "z'", "zre", "real", "re(z)/ohm", "z'/ohm", "z'_raw [ohm]"},
    "z_imag": {
        "zimag",
        "zimag_ohm",
        "z imag",
        "z''",
        "-zimag",
        "-z''",
        "imag",
        "im(z)/ohm",
        "-z''/ohm",
        "z\"_raw [ohm]",
    },
    "sheet_resistance": {"sheet resistance", "sheet_resistance", "resistance", "ohm/sq", "ohm per sq", "ohm/square"},
    "point": {"point", "position", "replicate", "spot"},
    "c_rate": {"c-rate", "c rate", "crate", "rate", "c_rate"},
}


def parse_file(path: Path) -> ParsedDataset:
    suffix = path.suffix.lower()
    warning = ""
    rows: list[dict[str, Any]]
    if suffix in {".csv", ".tsv", ".txt"}:
        rows = read_delimited(path)
    elif suffix in {".seo", ".sde"} and parse_wonatech_eis_file is not None:
        try:
            rows = read_wonatech_eis_binary(path)
        except Exception:
            if suffix == ".sde":
                rows, warning = read_sde_text_table(path)
            else:
                rows = read_delimited(path)
    elif suffix == ".sde":
        rows, warning = read_sde_text_table(path)
    elif suffix in {".xlsx", ".xls"}:
        rows = read_xlsx_optional(path)
    else:
        rows = read_delimited(path)
    normalized = normalize_rows(rows)
    analysis_type = detect_analysis_type(path.name, normalized)
    if analysis_type == ANALYSIS_VOLTAGE:
        normalized = expand_voltage_profile_rows(normalized)
    elif analysis_type == ANALYSIS_CAPACITY:
        normalized = normalize_capacity_rows(normalized)
    meta = build_file_meta(path, analysis_type, warning)
    if analysis_type == ANALYSIS_EIS:
        ensure_eis_fit_metadata(path, normalized)
    return ParsedDataset(meta=meta, rows=normalized, columns=list(normalized[0].keys()) if normalized else [])


def parse_eis_file(path: Path) -> dict[str, Any]:
    dataset = parse_file(path)
    return parsed_dataset_to_eis_arrays(dataset)


def parsed_dataset_to_eis_arrays(dataset: ParsedDataset) -> dict[str, Any]:
    z_real: list[float] = []
    z_imag: list[float] = []
    frequency: list[float | None] = []
    for row in dataset.rows:
        real = numeric_value(row.get("z_real"))
        imag = numeric_value(row.get("z_imag"))
        if real is None or imag is None:
            continue
        z_real.append(real)
        z_imag.append(imag)
        frequency.append(numeric_value(row.get("frequency")))
    return {
        "z_real": z_real,
        "z_imag": z_imag,
        "frequency": frequency,
        "source_format": dataset.meta.path.suffix.lower().lstrip(".") or "unknown",
    }


def ensure_eis_fit_metadata(path: Path, rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    if fit_eis_first_arc is None or load_valid_fit_metadata is None or save_fit_metadata is None:
        return None
    valid = load_valid_fit_metadata(path)
    if valid is not None:
        return valid
    z_real: list[float] = []
    z_imag: list[float] = []
    for row in rows:
        real = numeric_value(row.get("z_real"))
        imag = numeric_value(row.get("z_imag"))
        if real is None or imag is None:
            continue
        z_real.append(real)
        z_imag.append(imag)
    if not z_real:
        return None
    result = fit_eis_first_arc(z_real, z_imag)
    save_fit_metadata(
        path,
        result,
        extra={"source_format": path.suffix.lower().lstrip(".") or "unknown", "point_count": len(z_real)},
    )
    return load_valid_fit_metadata(path)


def numeric_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_delimited(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in sample else csv.excel
    raw_rows = list(csv.reader(text.splitlines(), dialect=dialect))
    if len(raw_rows) >= 3 and looks_like_voltage_profile_header(raw_rows[0], raw_rows[1]):
        headers = []
        for idx, first in enumerate(raw_rows[0]):
            second = raw_rows[1][idx] if idx < len(raw_rows[1]) else ""
            headers.append(f"{first.strip()} {second.strip()}".strip() or f"column_{idx + 1}")
        output = []
        for row in raw_rows[2:]:
            record = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)}
            if any(str(value).strip() for value in record.values()):
                output.append(record)
        return output
    if raw_rows and looks_like_data_row(raw_rows[0]):
        return rows_to_records(raw_rows)
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return [dict(row) for row in reader if any((value or "").strip() for value in row.values())]


def read_xlsx_optional(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:
        if sheet_name:
            raise ModuleNotFoundError("openpyxl is required when reading a named workbook sheet.") from exc
        return read_xlsx_builtin(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    output = []
    worksheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    for sheet in worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        output.extend(rows_to_records(rows))
    return output


def read_xlsx_builtin(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_names = sorted(name for name in archive.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", name))
        output: list[dict[str, Any]] = []
        for sheet_name in sheet_names:
            rows = read_sheet_xml(archive, sheet_name, shared_strings)
            output.extend(rows_to_records(rows))
        return output


def read_wonatech_eis_binary(path: Path) -> list[dict[str, Any]]:
    if parse_wonatech_eis_file is None:
        raise ValueError("WonATech EIS parser is unavailable.")
    result = parse_wonatech_eis_file(path)
    return [
        {
            "point": record.point,
            "frequency": record.frequency_hz,
            "z_real": record.zreal_ohm,
            "z_imag": record.zimag_ohm,
        }
        for record in result.records
    ]


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings = []
    for item in root.findall("main:si", ns):
        texts = [node.text or "" for node in item.findall(".//main:t", ns)]
        strings.append("".join(texts))
    return strings


def read_sheet_xml(archive: zipfile.ZipFile, sheet_name: str, shared_strings: list[str]) -> list[list[Any]]:
    root = ET.fromstring(archive.read(sheet_name))
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: list[list[Any]] = []
    for row in root.findall(".//main:row", ns):
        values: list[Any] = []
        for cell in row.findall("main:c", ns):
            col_idx = column_index(cell.attrib.get("r", "A1"))
            while len(values) < col_idx:
                values.append(None)
            values.append(cell_value(cell, shared_strings, ns))
        rows.append(values)
    return rows


def cell_value(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.findall(".//main:t", ns)]
        return "".join(texts)
    value = cell.find("main:v", ns)
    if value is None or value.text is None:
        return None
    text = value.text
    if cell_type == "s":
        idx = int(text)
        return shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return text


def column_index(cell_ref: str) -> int:
    letters = re.match(r"([A-Z]+)", cell_ref.upper())
    if not letters:
        return 1
    index = 0
    for char in letters.group(1):
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def rows_to_records(rows: list[tuple[Any, ...]] | list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    first_row_is_data = looks_like_data_row(rows[0])
    headers = (
        [f"column_{idx + 1}" for idx in range(len(rows[0]))]
        if first_row_is_data
        else [str(value).strip() if value is not None else f"column_{idx + 1}" for idx, value in enumerate(rows[0])]
    )
    output = []
    data_rows = rows if first_row_is_data else rows[1:]
    for row in data_rows:
        record = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)}
        if any(value not in (None, "") for value in record.values()):
            output.append(record)
    return output


def looks_like_data_row(row: tuple[Any, ...] | list[Any]) -> bool:
    values = [value for value in row if value not in (None, "")]
    if not values:
        return False
    numeric = sum(1 for value in values if is_number_like(value))
    text = " ".join(str(value).lower() for value in values)
    header_words = ("cycle", "capacity", "voltage", "z'", "sample", "binder", "전해질", "합제")
    return numeric >= max(2, len(values) // 2) and not any(word in text for word in header_words)


def is_number_like(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    return bool(re.fullmatch(r"\s*[-+]?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][-+]?\d+)?\s*", str(value)))


def read_sde_text_table(path: Path) -> tuple[list[dict[str, Any]], str]:
    raw = path.read_bytes()
    if raw and raw.count(b"\x00") / len(raw) > 0.08:
        return [], "Binary SDE file detected. Convert it through the WonATech/ZIVE parser service."
    text = raw.decode("utf-8", errors="ignore")
    if not text.strip():
        text = raw.decode("latin-1", errors="ignore")
    numeric_rows: list[list[str]] = []
    for line in text.splitlines():
        values = re.findall(r"[-+]?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][-+]?\d+)?", line)
        if len(values) >= 2:
            numeric_rows.append(values)
    if not numeric_rows:
        return [], "SDE file could not be decoded as a text-like numeric table."
    width = max(len(row) for row in numeric_rows)
    if width >= 3:
        headers = ["frequency", "z_real", "z_imag"] + [f"extra_{idx}" for idx in range(4, width + 1)]
    else:
        headers = ["z_real", "z_imag"]
    rows = []
    for values in numeric_rows:
        padded = values + [""] * (len(headers) - len(values))
        rows.append(dict(zip(headers, padded)))
    return rows, "Parsed SDE with best-effort numeric text extraction."


def normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for row in rows:
        output: dict[str, Any] = {}
        for key, value in row.items():
            canonical = canonical_column(str(key))
            output[canonical] = value
        normalized.append(output)
    return normalized


def canonical_column(name: str) -> str:
    cleaned = name.replace("\n", " ")
    cleaned = re.sub(r"\s+", " ", cleaned.strip().lower())
    cleaned = cleaned.replace("−", "-")
    cleaned = re.sub(r"^[a-z]\s+", "", cleaned)
    for canonical, aliases in ALIASES.items():
        if cleaned in aliases:
            return canonical
    if "z'_raw" in cleaned or "z′_raw" in cleaned or "z real" in cleaned:
        return "z_real"
    if 'z"_raw' in cleaned or "z''_raw" in cleaned or "z imag" in cleaned:
        return "z_imag"
    compact = re.sub(r"[^a-z0-9]+", "_", cleaned).strip("_")
    return compact or "column"


def detect_analysis_type(filename: str, rows: list[dict[str, Any]]) -> str:
    lowered = filename.lower()
    cols = set(rows[0].keys()) if rows else set()
    if any(token in lowered for token in ("eis", "sde", "nyquist")) or {"z_real", "z_imag"} <= cols:
        return ANALYSIS_EIS
    if "raman" in lowered:
        return ANALYSIS_RAMAN
    if "tga" in lowered or "thermogravimetric" in lowered:
        return ANALYSIS_TGA
    if "sheet" in lowered or "resistance" in lowered or "sheet_resistance" in cols:
        return ANALYSIS_SHEET
    if (
        "voltage" in lowered
        or "profile" in lowered
        or ("cycle" in lowered and any("1st" in col or "10th" in col for col in cols))
        or ({"capacity", "voltage"} <= cols and "direction" in cols)
    ):
        return ANALYSIS_VOLTAGE
    if "capacity" in lowered or {"cycle", "charge_capacity", "discharge_capacity"} <= cols or looks_like_capacity_columns(cols):
        return ANALYSIS_CAPACITY
    return ANALYSIS_UNKNOWN


def looks_like_capacity_columns(cols: set[str]) -> bool:
    return {"column_1", "column_4", "column_5"} <= cols


def build_file_meta(path: Path, analysis_type: str, warning: str = "") -> FileMeta:
    stem = path.stem
    parts = stem.split("__")
    if len(parts) >= 2:
        cell_id = parts[0]
        time_point = parts[2] if len(parts) >= 3 else ""
        date = parts[3] if len(parts) >= 4 else ""
    else:
        cell_id = guess_cell_id(stem, analysis_type)
        time_point = guess_time_point(stem)
        date = guess_date(stem)
    return FileMeta(
        path=path,
        original_filename=path.name,
        analysis_type=analysis_type,
        cell_id=cell_id,
        time_point=time_point,
        date=date,
        warning=warning,
    )


def guess_cell_id(stem: str, analysis_type: str) -> str:
    cleaned = re.sub(r"(?i)(capacity|cycle|voltage|profile|eis|nyquist|sheet|resistance|raman|tga|thermogravimetric)", "", stem)
    cleaned = re.sub(r"(?i)(?:^|[_\-\s])\d+hr(?:$|[_\-\s])", "_", cleaned)
    cleaned = re.sub(r"[_\-\s]+", "_", cleaned).strip("_")
    return cleaned or re.sub(r"\W+", "_", stem).strip("_") or "unknown_cell"


def guess_time_point(stem: str) -> str:
    match = re.search(r"(?i)(\d+\s*hr|\d+\s*h)", stem)
    return re.sub(r"\s+", "", match.group(1)).lower() if match else ""


def guess_date(stem: str) -> str:
    match = re.search(r"(20\d{6}|20\d{2}[-_]\d{2}[-_]\d{2})", stem)
    return match.group(1).replace("_", "-") if match else ""


def looks_like_voltage_profile_header(first: list[str], second: list[str]) -> bool:
    first_text = " ".join(first).lower()
    second_text = " ".join(second).lower()
    return bool(re.search(r"\b(1st|2nd|3rd|\d+th)\s+(ch|dis)", first_text)) and "v [v]" in second_text


def expand_voltage_profile_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return rows
    if {"cycle", "capacity", "voltage"} <= set(rows[0]):
        return rows
    groups: dict[tuple[str, str], dict[str, str]] = {}
    for column in rows[0]:
        match = re.match(r"(?P<cycle>\d+)(?:st|nd|rd|th)_+(?P<direction>ch|dis)_+(?P<field>.+)", column)
        if not match:
            continue
        cycle = match.group("cycle")
        direction = "charge" if match.group("direction") == "ch" else "discharge"
        field = match.group("field")
        key = (cycle, direction)
        groups.setdefault(key, {})
        if field.startswith("v"):
            groups[key]["voltage"] = column
        elif field.startswith("q_m") or "mah_g" in field:
            groups[key]["capacity"] = column
        elif field.startswith("q") and "capacity" not in groups[key]:
            groups[key]["capacity"] = column
    if not groups:
        return rows
    expanded: list[dict[str, Any]] = []
    for row in rows:
        for (cycle, direction), columns in sorted(groups.items(), key=lambda item: (int(item[0][0]), item[0][1])):
            voltage_col = columns.get("voltage")
            capacity_col = columns.get("capacity")
            if not voltage_col or not capacity_col:
                continue
            voltage = row.get(voltage_col)
            capacity = row.get(capacity_col)
            if voltage in (None, "") or capacity in (None, ""):
                continue
            expanded.append({"cycle": cycle, "direction": direction, "voltage": voltage, "capacity": capacity})
    return expanded or rows


def normalize_capacity_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return rows
    if {"cycle", "charge_capacity", "discharge_capacity"} <= set(rows[0]):
        return rows
    if not {"column_1", "column_4", "column_5"} <= set(rows[0]):
        return rows
    normalized = []
    for row in rows:
        output = dict(row)
        output.setdefault("cycle", row.get("column_1"))
        output.setdefault("charge_capacity", row.get("column_4"))
        output.setdefault("discharge_capacity", row.get("column_5"))
        normalized.append(output)
    return normalized
