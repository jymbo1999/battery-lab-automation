from __future__ import annotations

import hashlib
import csv
import json
import logging
import math
import re
import shutil
import time
import uuid
from dataclasses import asdict, dataclass, fields, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from .capacity_matching import write_capacity_match_outputs
from .conditions import CONDITION_FIELDS, condition_column, find_condition, read_conditions
from .eis_matching import write_eis_match_outputs
from .excel_dashboard import apply_row_formulas
from .file_io import parse_file
from .matching_service import EIS_SUFFIXES, collect_capacity_summary_sources, collect_source_files, load_match_overrides, save_match_overrides
from .metrics import compute_metrics
from .models import MetricRecord
from .plots import write_dataset_plot
from .wonatech_service import convert_wonatech_file, is_wonatech_source


logger = logging.getLogger("battery_lab.import")

ALLOWED_IMPORT_SUFFIXES = {".sde", ".seo", ".wrd", ".csv", ".xlsx", ".xls"}
ASSIGNMENT_LABELS = {
    "eis_comparison": "EIS 비교클러스터",
    "eis_time_series": "EIS 시계열",
    "capacity_1": "Capacity 1) 0.1C continuous",
    "capacity_2": "Capacity 2) 안정화 후 0.5C",
    "capacity_3": "Capacity 3) rate performance",
    "exclude": "이번 등록에서 제외",
}
ALL_ASSIGNMENT_OPTIONS = list(ASSIGNMENT_LABELS)
# Options offered in the wizard's per-file type toggle. "exclude" is intentionally
# omitted here: removing a file is done with the row's × delete button, not a toggle
# value. "exclude" remains a valid internal assignment (commit still skips it).
TYPE_OPTIONS = ["eis_comparison", "eis_time_series", "capacity_1", "capacity_2", "capacity_3"]
REQUIRED_IMPORT_FIELDS = ["date", "sample", "foil_electrode_g", "foil_electrode_mm"]
NUMERIC_IMPORT_FIELDS = ["foil_electrode_g", "foil_electrode_mm", "foil_g", "ratio", "current_density", "foil_thickness_mm", "electrolyte_ul"]
# Single source of truth for the import form + journal writer.
# Each field: stable key, EXACT Excel header (sheet JYJ), bucket, default.
# Mapping by EXACT header avoids the condition_column 'mm' collision
# (호일 두께/전극 두께/압연 전 두께 all normalize to 'mm').
IMPORT_JOURNAL_FIELDS = [
    {"key": "date", "header": "Date", "bucket": "variable", "default": ""},
    {"key": "sample", "header": "Sample", "bucket": "variable", "default": ""},
    {"key": "foil_electrode_g", "header": "foil+electrode (g)", "bucket": "variable", "default": ""},
    {"key": "foil_electrode_mm", "header": "전극(foil+electrode) 두께(mm)", "bucket": "variable", "default": ""},
    {"key": "reference", "header": "참고", "bucket": "fixed", "default": "12 파이_Cu foil"},
    {"key": "electrolyte", "header": "전해질", "bucket": "fixed", "default": "1.0M LiPF6 EC/DEC 1:1"},
    {"key": "cell_type", "header": "종류", "bucket": "fixed", "default": "LIB"},
    {"key": "conductive_agent", "header": "Conductive agent", "bucket": "fixed", "default": "-"},
    {"key": "binder", "header": "Binder", "bucket": "fixed", "default": "2wt%cmc"},
    {"key": "voltage_range", "header": "Voltage range", "bucket": "fixed", "default": "0.01~2V"},
    {"key": "foil_g", "header": "foil (g)", "bucket": "fixed", "default": "0.009928"},
    {"key": "ratio", "header": "ratio", "bucket": "fixed", "default": "0.96"},
    {"key": "current_density", "header": "Current density (mA/g)", "bucket": "fixed", "default": "37.2"},
    {"key": "foil_thickness_mm", "header": "호일 두께(mm)", "bucket": "fixed", "default": "0.00958"},
    {"key": "electrolyte_ul", "header": "Electrolyte (ul)", "bucket": "fixed", "default": "80"},
    {"key": "drying_condition", "header": "Drying Condition", "bucket": "fixed", "default": "60도 12시간"},
]
# Binder presets offered in the form dropdown (still free-text editable).
BINDER_PRESETS = ["2wt%cmc", "2wt%cmc/40wt%SBR"]


def field_keys() -> list[str]:
    return [f["key"] for f in IMPORT_JOURNAL_FIELDS]


def variable_keys() -> list[str]:
    return [f["key"] for f in IMPORT_JOURNAL_FIELDS if f["bucket"] == "variable"]


def fixed_defaults() -> dict[str, str]:
    return {f["key"]: f["default"] for f in IMPORT_JOURNAL_FIELDS if f["bucket"] == "fixed"}


def _to_float(value: object) -> float | None:
    text = str(value if value is not None else "").replace(",", "").strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def compute_derived_metadata(metadata: dict[str, object]) -> dict[str, float | None]:
    """Compute the journal's formula-column values in Python so the app can read
    them numerically before Excel ever recalculates (read_xlsx_optional uses
    data_only=True). Mirrors excel_dashboard.FORMULA_TEMPLATES_BY_HEADER."""
    foil_electrode_g = _to_float(metadata.get("foil_electrode_g"))
    foil_g = _to_float(metadata.get("foil_g"))
    ratio = _to_float(metadata.get("ratio"))
    foil_electrode_mm = _to_float(metadata.get("foil_electrode_mm"))
    foil_thickness_mm = _to_float(metadata.get("foil_thickness_mm"))

    active = None
    areal = None
    if None not in (foil_electrode_g, foil_g, ratio):
        active = (foil_electrode_g - foil_g) * ratio
        areal = active * 1000 / (math.pi * 0.6 ** 2)

    electrode_g = None
    electrode_density = None
    if None not in (foil_electrode_g, foil_g):
        electrode_g = foil_electrode_g - foil_g
        if None not in (foil_electrode_mm, foil_thickness_mm):
            thickness = foil_electrode_mm - foil_thickness_mm
            volume = 113.1 * thickness
            if volume:
                electrode_density = electrode_g / (volume / 1000)

    return {
        "active_material_g": active,
        "areal_mass_density": areal,
        "electrode_g": electrode_g,
        "electrode_density": electrode_density,
    }


EIS_CLUSTER_FIELDS = ("electrolyte", "binder", "voltage_range", "ratio")
CAPACITY_CLUSTER_FIELDS = ("cell_type", "electrolyte", "binder", "voltage_range", "ratio")


@dataclass(frozen=True)
class DraftImportFile:
    file_id: str
    original_filename: str
    raw_path: str
    suffix: str
    sha256: str
    size_bytes: int
    parser_kind: str
    analysis_type: str
    cell_id: str
    normalized_rows: int
    processed_path: str = ""
    parser_meta_path: str = ""
    raw_timeseries_path: str = ""
    plot_path: str = ""
    plot_meta_path: str = ""
    warning: str = ""
    metrics: dict[str, object] | None = None
    time_point: str = ""
    suggested_assignment: str = ""
    assignment: str = ""
    assignment_options: list[str] | None = None
    assignment_reason: str = ""
    auto_assignments: list[str] | None = None


@dataclass(frozen=True)
class DraftImportManifest:
    draft_id: str
    created_at: str
    draft_root: str
    raw_dir: str
    processed_dir: str
    preview_dir: str
    files: list[DraftImportFile]
    errors: list[str]
    updated_at: str = ""
    metadata: dict[str, object] | None = None
    metadata_status: str = "missing"
    metadata_errors: list[str] | None = None
    commit_status: str = "draft"
    committed_at: str = ""
    journal_row: int | None = None
    saved_files: list[dict[str, object]] | None = None
    match_overrides: list[dict[str, object]] | None = None
    persist_outputs: list[dict[str, object]] | None = None


def create_import_draft(
    uploads: list[tuple[str, BinaryIO]],
    output_root: Path,
    *,
    draft_id: str | None = None,
    write_raw_wrd: bool = False,
) -> DraftImportManifest:
    draft_id = draft_id or uuid.uuid4().hex
    draft_root = output_root / "import_drafts" / draft_id
    raw_dir = draft_root / "raw"
    processed_dir = draft_root / "processed"
    preview_dir = draft_root / "preview"
    raw_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)
    preview_dir.mkdir(parents=True, exist_ok=True)

    files, errors = _ingest_uploads(
        uploads,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
        preview_dir=preview_dir,
        used_names=set(),
        write_raw_wrd=write_raw_wrd,
    )

    manifest = DraftImportManifest(
        draft_id=draft_id,
        created_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        draft_root=str(draft_root),
        raw_dir=str(raw_dir),
        processed_dir=str(processed_dir),
        preview_dir=str(preview_dir),
        files=files,
        errors=errors,
        updated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        metadata={},
        metadata_status="missing",
        metadata_errors=[],
        commit_status="draft",
        saved_files=[],
        match_overrides=[],
        persist_outputs=[],
    )
    write_manifest(manifest, draft_root / "manifest.json")
    return manifest


def _ingest_uploads(
    uploads: list[tuple[str, BinaryIO]],
    *,
    raw_dir: Path,
    processed_dir: Path,
    preview_dir: Path,
    used_names: set[str],
    write_raw_wrd: bool,
) -> tuple[list[DraftImportFile], list[str]]:
    files: list[DraftImportFile] = []
    errors: list[str] = []
    for original_filename, stream in uploads:
        safe_name = unique_filename(safe_filename(original_filename), used_names)
        suffix = Path(safe_name).suffix.lower()
        if suffix not in ALLOWED_IMPORT_SUFFIXES:
            errors.append(f"{original_filename}: unsupported extension {suffix or '(none)'}")
            continue
        raw_path = raw_dir / safe_name
        with raw_path.open("wb") as handle:
            shutil.copyfileobj(stream, handle)
        try:
            files.append(
                build_draft_file(
                    raw_path,
                    original_filename=original_filename,
                    processed_dir=processed_dir,
                    preview_dir=preview_dir,
                    write_raw_wrd=write_raw_wrd,
                )
            )
        except Exception as exc:
            errors.append(f"{original_filename}: {exc}")
    return files, errors


def append_import_draft_files(
    output_root: Path,
    draft_id: str,
    uploads: list[tuple[str, BinaryIO]],
    *,
    write_raw_wrd: bool = False,
) -> DraftImportManifest:
    """Parse and append more uploads into an existing (uncommitted) draft."""
    manifest = load_import_draft(output_root, draft_id)
    if manifest.commit_status == "committed":
        raise ValueError("Committed drafts cannot accept more files.")
    raw_dir = Path(manifest.raw_dir)
    processed_dir = Path(manifest.processed_dir)
    preview_dir = Path(manifest.preview_dir)
    for directory in (raw_dir, processed_dir, preview_dir):
        directory.mkdir(parents=True, exist_ok=True)
    used_names = {Path(item.raw_path).name for item in manifest.files}
    new_files, errors = _ingest_uploads(
        uploads,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
        preview_dir=preview_dir,
        used_names=used_names,
        write_raw_wrd=write_raw_wrd,
    )
    updated = replace(
        manifest,
        files=list(manifest.files) + new_files,
        errors=errors,
        updated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    write_manifest(updated, output_root / "import_drafts" / draft_id / "manifest.json")
    return updated


def remove_import_draft_file(output_root: Path, draft_id: str, file_id: str) -> DraftImportManifest:
    """Remove one file from an (uncommitted) draft and unlink its artifacts."""
    manifest = load_import_draft(output_root, draft_id)
    if manifest.commit_status == "committed":
        raise ValueError("Committed drafts cannot be edited.")
    target = next((item for item in manifest.files if item.file_id == file_id), None)
    if target is None:
        raise ValueError(f"file_id {file_id!r} not found in draft")
    for path_str in (
        target.raw_path,
        target.processed_path,
        target.parser_meta_path,
        target.raw_timeseries_path,
        target.plot_path,
        target.plot_meta_path,
    ):
        if not path_str:
            continue
        try:
            Path(path_str).unlink(missing_ok=True)
        except OSError:
            pass
    updated = replace(
        manifest,
        files=[item for item in manifest.files if item.file_id != file_id],
        updated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    write_manifest(updated, output_root / "import_drafts" / draft_id / "manifest.json")
    return updated


def build_draft_file(
    raw_path: Path,
    *,
    original_filename: str,
    processed_dir: Path,
    preview_dir: Path,
    write_raw_wrd: bool = False,
) -> DraftImportFile:
    parse_path = raw_path
    parser_kind = "table"
    processed_path = ""
    parser_meta_path = ""
    raw_timeseries_path = ""

    timings: dict[str, float] = {}
    t_start = time.perf_counter()

    def _mark(stage: str, since: float) -> float:
        now = time.perf_counter()
        timings[stage] = now - since
        return now

    t_cursor = t_start

    if is_wonatech_source(raw_path):
        # WRD -> capacity summary conversion.
        #
        # mAh/g normalization mapping (see wonatech_parsers.wrd.build_capacity_summary):
        #   raw Ah * 1000 = mAh  ->  mAh / mass_g = mAh/g
        #   mass_g = areal_mass_density(mg/cm^2) * (pi * 0.6**2 cm^2) / 1000
        #
        # mass_g is intentionally NOT passed here: at upload/draft time the
        # experiment info (areal_mass_density) has not been entered yet, so only
        # absolute mAh can be produced. The areal_mass_density arrives later via
        # the experiment-info form (see validate_metadata) and the capacity
        # rebuild job (routes.queue_import_rebuild_jobs).
        #
        # This gap was historically hidden because the app read pre-normalized
        # official "_Capacity.csv" files (already mAh/g). For the WRD-only upload
        # flow, mass_g is threaded through to convert_wonatech_file(...) once the
        # condition is known, in battery_lab.ui.build_analysis_artifacts (the
        # build_capacity_graphs rebuild path), so mAh/g reproduces the legacy CSV.
        conversion = convert_wonatech_file(raw_path, processed_dir, write_raw_wrd=write_raw_wrd)
        parse_path = conversion.primary_csv_path
        parser_kind = conversion.kind
        processed_path = str(conversion.primary_csv_path)
        parser_meta_path = str(conversion.meta_path)
        raw_timeseries_path = str(conversion.raw_csv_path) if conversion.raw_csv_path else ""
        t_cursor = _mark("wonatech_convert", t_cursor)

    dataset = parse_file(parse_path)
    t_cursor = _mark("parse_file", t_cursor)
    record = compute_metrics(dataset)
    assignment = infer_assignment(dataset.meta.analysis_type, record.metrics, dataset.meta.time_point, raw_path.name)
    t_cursor = _mark("metrics_assign", t_cursor)
    plot_path = write_dataset_plot(dataset, preview_dir)
    plot_meta_path = plot_path.with_name(plot_path.name + ".meta.json") if plot_path else None
    t_cursor = _mark("write_plot", t_cursor)
    stat = raw_path.stat()
    sha256 = sha256_file(raw_path)
    t_cursor = _mark("sha256", t_cursor)

    total = time.perf_counter() - t_start
    logger.warning(
        "[IMPORT_TIME] file=%s type=%s size=%dB total=%.2fs %s",
        raw_path.name,
        dataset.meta.analysis_type,
        int(stat.st_size),
        total,
        " ".join(f"{stage}={dt:.2f}s" for stage, dt in timings.items()),
    )

    return DraftImportFile(
        file_id=f"{safe_filename(raw_path.stem)}__{sha256[:12]}",
        original_filename=original_filename,
        raw_path=str(raw_path),
        suffix=raw_path.suffix.lower(),
        sha256=sha256,
        size_bytes=int(stat.st_size),
        parser_kind=parser_kind,
        analysis_type=dataset.meta.analysis_type,
        cell_id=dataset.meta.cell_id,
        normalized_rows=len(dataset.rows),
        processed_path=processed_path,
        parser_meta_path=parser_meta_path,
        raw_timeseries_path=raw_timeseries_path,
        plot_path=str(plot_path) if plot_path else "",
        plot_meta_path=str(plot_meta_path) if plot_meta_path and plot_meta_path.exists() else "",
        warning=record.warning or dataset.meta.warning,
        metrics=record.metrics,
        time_point=dataset.meta.time_point,
        suggested_assignment=assignment["suggested_assignment"],
        assignment=assignment["suggested_assignment"],
        assignment_options=assignment["assignment_options"],
        assignment_reason=assignment["assignment_reason"],
        auto_assignments=assignment["auto_assignments"],
    )


def write_manifest(manifest: DraftImportManifest, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(asdict(manifest), ensure_ascii=False, indent=2), encoding="utf-8")


def load_import_draft(output_root: Path, draft_id: str) -> DraftImportManifest:
    path = output_root / "import_drafts" / draft_id / "manifest.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    return manifest_from_payload(payload)


def update_import_draft_assignments(
    output_root: Path,
    draft_id: str,
    assignments: dict[str, str],
) -> DraftImportManifest:
    manifest = load_import_draft(output_root, draft_id)
    updated_files = []
    for item in manifest.files:
        requested = assignments.get(item.file_id)
        if not requested:
            updated_files.append(item)
            continue
        options = item.assignment_options or ALL_ASSIGNMENT_OPTIONS
        if requested not in options:
            raise ValueError(f"{item.original_filename}: assignment {requested!r} is not allowed")
        updated_files.append(replace(item, assignment=requested))
    updated = replace(
        manifest,
        files=updated_files,
        updated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    write_manifest(updated, output_root / "import_drafts" / draft_id / "manifest.json")
    return updated


def update_import_draft_metadata(
    output_root: Path,
    draft_id: str,
    metadata: dict[str, object],
) -> DraftImportManifest:
    manifest = load_import_draft(output_root, draft_id)
    cleaned = clean_metadata(metadata)
    errors = validate_metadata(cleaned)
    updated = replace(
        manifest,
        metadata=cleaned,
        metadata_status="ready" if not errors else "invalid",
        metadata_errors=errors,
        updated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    write_manifest(updated, output_root / "import_drafts" / draft_id / "manifest.json")
    return updated


def commit_import_draft(
    output_root: Path,
    draft_id: str,
    *,
    eis_root: Path,
    capacity_root: Path,
    condition_workbook: Path,
    condition_sheet: str,
    eis_match_override_path: Path | None = None,
    capacity_match_override_path: Path | None = None,
) -> DraftImportManifest:
    manifest = load_import_draft(output_root, draft_id)
    if manifest.commit_status == "committed":
        return manifest
    if manifest.metadata_status != "ready":
        raise ValueError("Draft metadata must be ready before commit.")
    active_files = [item for item in manifest.files if item.assignment != "exclude"]
    if not active_files:
        raise ValueError("No files selected for commit.")

    journal_row = append_journal_row(condition_workbook, condition_sheet, manifest.metadata or {})
    saved_files = save_draft_files_to_final_locations(
        manifest,
        journal_row=journal_row,
        eis_root=eis_root,
        capacity_root=capacity_root,
        output_root=output_root,
    )
    match_overrides = write_commit_match_overrides(
        manifest,
        journal_row=journal_row,
        saved_files=saved_files,
        eis_root=eis_root,
        capacity_root=capacity_root,
        condition_workbook=condition_workbook,
        condition_sheet=condition_sheet,
        eis_match_override_path=eis_match_override_path,
        capacity_match_override_path=capacity_match_override_path,
    )
    persist_outputs = persist_commit_outputs(
        manifest,
        saved_files=saved_files,
        output_root=output_root,
        eis_root=eis_root,
        capacity_root=capacity_root,
        condition_workbook=condition_workbook,
        condition_sheet=condition_sheet,
        eis_match_override_path=eis_match_override_path,
        capacity_match_override_path=capacity_match_override_path,
    )
    committed = replace(
        manifest,
        commit_status="committed",
        committed_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        journal_row=journal_row,
        saved_files=saved_files,
        match_overrides=match_overrides,
        persist_outputs=persist_outputs,
        updated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )
    write_manifest(committed, output_root / "import_drafts" / draft_id / "manifest.json")
    return committed


def persist_commit_outputs(
    manifest: DraftImportManifest,
    *,
    saved_files: list[dict[str, object]],
    output_root: Path,
    eis_root: Path,
    capacity_root: Path,
    condition_workbook: Path,
    condition_sheet: str,
    eis_match_override_path: Path | None,
    capacity_match_override_path: Path | None,
) -> list[dict[str, object]]:
    results: list[dict[str, object]] = []
    try:
        conditions = read_conditions(condition_workbook, sheet_name=condition_sheet) if condition_workbook.exists() else {}
    except Exception as exc:
        conditions = {}
        results.append({"kind": "conditions", "ok": False, "error": str(exc), "path": str(condition_workbook)})

    datasets = []
    records: list[MetricRecord] = []
    for saved in saved_files:
        kind = "capacity" if str(saved.get("assignment") or "").startswith("capacity_") else "eis"
        source_path = override_source_path_for_saved_file(kind, saved)
        if not source_path:
            continue
        try:
            dataset = parse_file(Path(source_path))
            record = compute_metrics(dataset)
            plot_path = write_dataset_plot(dataset, output_root)
            datasets.append(dataset)
            records.append(record)
            results.append(
                {
                    "kind": "plot",
                    "analysis_type": dataset.meta.analysis_type,
                    "ok": True,
                    "source_path": str(source_path),
                    "artifact_path": str(plot_path) if plot_path else "",
                }
            )
        except Exception as exc:
            results.append({"kind": "plot", "ok": False, "source_path": str(source_path), "error": str(exc)})

    if records:
        try:
            upsert_summary_metrics(records, output_root / "summary_metrics.csv", conditions)
            results.append({"kind": "summary_metrics", "ok": True, "path": str(output_root / "summary_metrics.csv"), "records": len(records)})
        except Exception as exc:
            results.append({"kind": "summary_metrics", "ok": False, "path": str(output_root / "summary_metrics.csv"), "error": str(exc)})

    affected = {"capacity" if str(saved.get("assignment") or "").startswith("capacity_") else "eis" for saved in saved_files}
    if "eis" in affected:
        try:
            overrides = load_match_overrides(eis_match_override_path) if eis_match_override_path else {}
            source_paths = collect_source_files(eis_root, EIS_SUFFIXES)
            report = write_eis_match_outputs(source_paths, conditions, output_root, eis_root, overrides)
            results.append({"kind": "eis_match_outputs", "ok": True, "path": str(output_root / "eis_match_report.json"), "matches": len(report.matches)})
        except Exception as exc:
            results.append({"kind": "eis_match_outputs", "ok": False, "path": str(output_root / "eis_match_report.json"), "error": str(exc)})
    if "capacity" in affected:
        try:
            overrides = load_match_overrides(capacity_match_override_path) if capacity_match_override_path else {}
            source_paths = collect_capacity_summary_sources(capacity_root)
            report = write_capacity_match_outputs(source_paths, conditions, output_root, capacity_root, overrides)
            results.append({"kind": "capacity_match_outputs", "ok": True, "path": str(output_root / "capacity_match_report.json"), "matches": len(report.matches)})
        except Exception as exc:
            results.append({"kind": "capacity_match_outputs", "ok": False, "path": str(output_root / "capacity_match_report.json"), "error": str(exc)})
    return results


def upsert_summary_metrics(records: list[MetricRecord], path: Path, conditions: dict[str, dict[str, object]]) -> None:
    existing: list[dict[str, object]] = []
    if path.exists():
        with path.open(newline="", encoding="utf-8") as handle:
            existing = list(csv.DictReader(handle))
    new_rows = [summary_metric_row(record, conditions) for record in records]
    by_key = {summary_metric_key(row): row for row in existing}
    for row in new_rows:
        by_key[summary_metric_key(row)] = row
    rows = list(by_key.values())
    headers: list[str] = []
    for preferred in ["cell_id", "analysis_type", "source_file", "warning"]:
        if any(preferred in row for row in rows):
            headers.append(preferred)
    condition_keys = [key for key in CONDITION_FIELDS if any(key in row for row in rows)]
    metric_keys = sorted({key for row in rows for key in row if key not in set(headers) | set(condition_keys)})
    headers = headers + [key for key in condition_keys if key not in headers] + [key for key in metric_keys if key not in headers]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def summary_metric_row(record: MetricRecord, conditions: dict[str, dict[str, object]]) -> dict[str, object]:
    condition = find_condition(record.cell_id, conditions)
    row: dict[str, object] = {
        "cell_id": record.cell_id,
        "analysis_type": record.analysis_type,
        "source_file": record.source_file,
        "warning": record.warning,
    }
    row.update({key: condition.get(key, "") for key in CONDITION_FIELDS if key in condition})
    row.update(record.metrics)
    return row


def summary_metric_key(row: dict[str, object]) -> tuple[str, str, str]:
    return (str(row.get("analysis_type") or ""), str(row.get("source_file") or ""), str(row.get("cell_id") or ""))


def header_column_map(worksheet) -> dict[str, int]:
    """Map EXACT header text -> column index (1-based)."""
    out: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=col).value
        if header not in (None, ""):
            out[str(header).strip()] = col
    return out


def column_by_condition_key(worksheet, key: str) -> int | None:
    for col in range(1, worksheet.max_column + 1):
        if condition_column(worksheet.cell(row=1, column=col).value) == key:
            return col
    return None


def write_journal_row(worksheet, row: int, metadata: dict[str, object]) -> None:
    """Write one journal row by EXACT header, apply display formulas, then
    overwrite the app-read columns (areal_mass_density, electrode_density)
    with Python literals so data_only reads return numbers immediately."""
    by_header = header_column_map(worksheet)
    for field in IMPORT_JOURNAL_FIELDS:
        value = metadata.get(field["key"])
        if value in (None, ""):
            continue
        col = by_header.get(field["header"])
        if col:
            worksheet.cell(row=row, column=col).value = value
    apply_row_formulas(worksheet, row)
    derived = compute_derived_metadata(metadata)
    for key in ("areal_mass_density", "electrode_density"):
        col = column_by_condition_key(worksheet, key)
        if col and derived.get(key) is not None:
            worksheet.cell(row=row, column=col).value = derived[key]


def append_journal_row(condition_workbook: Path, condition_sheet: str, metadata: dict[str, object]) -> int:
    condition_workbook.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(condition_workbook)
    if condition_sheet not in workbook.sheetnames:
        workbook.close()
        raise KeyError(f"Sheet not found: {condition_sheet}")
    worksheet = workbook[condition_sheet]
    row = worksheet.max_row + 1
    write_journal_row(worksheet, row, metadata)
    workbook.save(condition_workbook)
    workbook.close()
    return row


def save_draft_files_to_final_locations(
    manifest: DraftImportManifest,
    *,
    journal_row: int,
    eis_root: Path,
    capacity_root: Path,
    output_root: Path,
) -> list[dict[str, object]]:
    saved = []
    metadata = manifest.metadata or {}
    for item in manifest.files:
        if item.assignment == "exclude":
            continue
        destination_dir = final_directory_for_item(item, metadata, journal_row, eis_root, capacity_root)
        destination_dir.mkdir(parents=True, exist_ok=True)
        raw_source = Path(item.raw_path)
        raw_target = collision_safe_path(destination_dir / final_filename_for_item(item, metadata, journal_row, raw_source.suffix))
        shutil.copy2(raw_source, raw_target)
        row = {
            "file_id": item.file_id,
            "assignment": item.assignment,
            "source_path": str(raw_source),
            "saved_path": str(raw_target),
            "processed_saved_path": "",
        }
        if item.processed_path:
            processed_source = Path(item.processed_path)
            processed_target = collision_safe_path(destination_dir / f"{raw_target.stem}_{processed_source.name}")
            shutil.copy2(processed_source, processed_target)
            row["processed_saved_path"] = str(processed_target)
        saved.append(row)

    processed_manifest_dir = output_root / "processed" / "imports" / manifest.draft_id
    processed_manifest_dir.mkdir(parents=True, exist_ok=True)
    write_manifest(manifest, processed_manifest_dir / "draft_manifest_before_commit.json")
    return saved


def write_commit_match_overrides(
    manifest: DraftImportManifest,
    *,
    journal_row: int,
    saved_files: list[dict[str, object]],
    eis_root: Path,
    capacity_root: Path,
    condition_workbook: Path,
    condition_sheet: str,
    eis_match_override_path: Path | None,
    capacity_match_override_path: Path | None,
) -> list[dict[str, object]]:
    if not eis_match_override_path and not capacity_match_override_path:
        return []

    condition_key, condition = condition_for_journal_row(condition_workbook, condition_sheet, journal_row, manifest.metadata or {})
    by_file = {item.file_id: item for item in manifest.files}
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    grouped: dict[str, tuple[Path, dict[str, dict[str, object]]]] = {}
    written: list[dict[str, object]] = []

    for saved in saved_files:
        item = by_file.get(str(saved.get("file_id") or ""))
        if item is None:
            continue
        kind = "capacity" if str(saved.get("assignment") or item.assignment).startswith("capacity_") else "eis"
        override_path = capacity_match_override_path if kind == "capacity" else eis_match_override_path
        source_root = capacity_root if kind == "capacity" else eis_root
        if override_path is None:
            continue
        source_path = override_source_path_for_saved_file(kind, saved)
        if not source_path:
            continue
        relative = relative_to_root(Path(source_path), source_root)
        if not relative:
            continue
        if kind not in grouped:
            grouped[kind] = (override_path, load_match_overrides(override_path))
        grouped[kind][1][relative] = {
            "condition_key": condition_key,
            "journal_row": journal_row,
            "sample": condition.get("sample") or (manifest.metadata or {}).get("sample") or condition_key,
            "date": condition.get("date") or (manifest.metadata or {}).get("date") or "",
            "selected_at": now,
            "selection_source": "import_commit",
            "import_draft_id": manifest.draft_id,
            "assignment": saved.get("assignment") or item.assignment,
        }
        written.append(
            {
                "kind": kind,
                "file_id": item.file_id,
                "relative_path": relative,
                "condition_key": condition_key,
                "journal_row": journal_row,
                "override_path": str(override_path),
            }
        )

    for override_path, overrides in grouped.values():
        save_match_overrides(override_path, overrides)
    return written


def condition_for_journal_row(
    condition_workbook: Path,
    condition_sheet: str,
    journal_row: int,
    metadata: dict[str, object],
) -> tuple[str, dict[str, object]]:
    conditions = read_conditions(condition_workbook, sheet_name=condition_sheet) if condition_workbook.exists() else {}
    for key, condition in conditions.items():
        if condition.get("_source_row_number") == journal_row:
            return key, condition
    fallback_key = str(metadata.get("sample") or f"row{journal_row}").strip() or f"row{journal_row}"
    return fallback_key, dict(metadata)


def override_source_path_for_saved_file(kind: str, saved: dict[str, object]) -> str:
    if kind == "capacity":
        processed = str(saved.get("processed_saved_path") or "")
        if processed:
            return processed
    return str(saved.get("saved_path") or "")


def relative_to_root(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except ValueError:
        return ""


# Human protocol tokens so capacity_matching.capacity_protocol_from_filename
# (looks for "rate per" / "0.5c" / "0.1c") and the data browser recognize
# imported capacity files exactly like legacy folders.
CAPACITY_PROTOCOL_TOKENS = {"capacity_1": "0.1C", "capacity_2": "0.5C", "capacity_3": "rate per"}


def assignment_protocol_token(assignment: str) -> str:
    return CAPACITY_PROTOCOL_TOKENS.get(assignment, assignment)


def final_directory_for_item(item: DraftImportFile, metadata: dict[str, object], journal_row: int, eis_root: Path, capacity_root: Path) -> Path:
    yymmdd = compact_metadata_date(metadata.get("date"))
    sample = safe_stem(str(metadata.get("sample") or item.cell_id or "sample"))
    if item.assignment.startswith("capacity_"):
        protocol = assignment_protocol_token(item.assignment)
        folder = safe_stem(f"{journal_row}_{sample}_{protocol}_cyc")
        return capacity_root / yymmdd / folder / long_metadata_date(metadata.get("date"))
    return eis_root / yymmdd / sample


def final_filename_for_item(item: DraftImportFile, metadata: dict[str, object], journal_row: int, suffix: str) -> str:
    sample = safe_stem(str(metadata.get("sample") or item.cell_id or "sample"))
    if item.assignment.startswith("capacity_"):
        token = assignment_protocol_token(item.assignment)
    else:
        token = safe_stem(item.time_point) if item.time_point else item.assignment
    return safe_filename(f"{journal_row}_{sample}_{token}{suffix.lower()}")


def collision_safe_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for idx in range(2, 1000):
        candidate = path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Too many filename collisions for {path}")


def compact_metadata_date(value: object) -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8 and digits.startswith("20"):
        return digits[2:]
    if len(digits) == 6:
        return digits
    return "unknown_date"


def long_metadata_date(value: object) -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        return digits
    if len(digits) == 6:
        return f"20{digits}"
    return "unknown_date"


def safe_stem(value: str) -> str:
    cleaned = re.sub(r"\s+", "_", value.strip())
    cleaned = re.sub(r"[^A-Za-z0-9_.가-힣-]+", "_", cleaned)
    return cleaned.strip("._") or "sample"


def manifest_from_payload(payload: dict[str, object]) -> DraftImportManifest:
    file_keys = {field.name for field in fields(DraftImportFile)}
    manifest_keys = {field.name for field in fields(DraftImportManifest)}
    files = [DraftImportFile(**{key: value for key, value in item.items() if key in file_keys}) for item in payload.get("files", [])]
    data = {key: value for key, value in payload.items() if key in manifest_keys and key != "files"}
    data["files"] = files
    data.setdefault("updated_at", data.get("created_at", ""))
    data.setdefault("metadata", {})
    data.setdefault("metadata_status", "missing")
    data.setdefault("metadata_errors", [])
    data.setdefault("commit_status", "draft")
    data.setdefault("committed_at", "")
    data.setdefault("journal_row", None)
    data.setdefault("saved_files", [])
    data.setdefault("match_overrides", [])
    data.setdefault("persist_outputs", [])
    return DraftImportManifest(**data)


def manifest_payload(manifest: DraftImportManifest) -> dict[str, object]:
    return asdict(manifest)


def clean_metadata(metadata: dict[str, object]) -> dict[str, object]:
    allowed = set(field_keys())
    cleaned: dict[str, object] = {}
    for key in allowed:
        value = metadata.get(key)
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
        if value != "":
            cleaned[key] = value
    return cleaned


def validate_metadata(metadata: dict[str, object]) -> list[str]:
    errors = [f"{field} is required" for field in REQUIRED_IMPORT_FIELDS if metadata.get(field) in (None, "")]
    for field in NUMERIC_IMPORT_FIELDS:
        value = metadata.get(field)
        if value in (None, ""):
            continue
        if _to_float(value) is None:
            errors.append(f"{field} must be numeric")
    fe = _to_float(metadata.get("foil_electrode_g"))
    foil = _to_float(metadata.get("foil_g"))
    if fe is not None and foil is not None and fe <= foil:
        errors.append("foil+electrode (g) must be greater than foil (g)")
    ratio = _to_float(metadata.get("ratio"))
    if ratio is not None and not (0 < ratio <= 1):
        errors.append("ratio must be between 0 and 1")
    date = str(metadata.get("date") or "")
    if date and not re.fullmatch(r"(?:\d{6}|\d{8}|\d{4}[-./]\d{1,2}[-./]\d{1,2})", date):
        errors.append("date must be YYMMDD, YYYYMMDD, or YYYY-MM-DD")
    return errors


def metadata_options_from_conditions(conditions: dict[str, dict[str, object]], *, limit: int = 40) -> dict[str, list[str]]:
    fields = REQUIRED_IMPORT_FIELDS + ["sample_group", "material_family", "treatment"]
    options: dict[str, list[str]] = {}
    for field in fields:
        values = sorted({str(row.get(field)).strip() for row in conditions.values() if row.get(field) not in (None, "")})
        options[field] = values[:limit]
    return options


def build_import_draft_cluster_preview(
    output_root: Path,
    draft_id: str,
    conditions: dict[str, dict[str, object]],
) -> dict[str, object]:
    manifest = load_import_draft(output_root, draft_id)
    metadata = manifest.metadata or {}
    rows = []
    for item in manifest.files:
        rows.append(cluster_preview_row(item, metadata, manifest.metadata_status, conditions))
    summary: dict[str, int] = {}
    for row in rows:
        summary[row["status"]] = summary.get(row["status"], 0) + 1
    return {
        "draft_id": manifest.draft_id,
        "metadata_status": manifest.metadata_status,
        "metadata_errors": manifest.metadata_errors or [],
        "rows": rows,
        "summary": summary,
    }


def cluster_preview_row(
    item: DraftImportFile,
    metadata: dict[str, object],
    metadata_status: str,
    conditions: dict[str, dict[str, object]],
) -> dict[str, object]:
    assignment = item.assignment or item.suggested_assignment
    if assignment == "exclude":
        return base_cluster_preview_row(item, assignment, "excluded", "이번 등록에서 제외됩니다.", [], "")
    if metadata_status != "ready":
        return base_cluster_preview_row(item, assignment, "metadata_required", "실험정보 저장 후 cluster preview가 가능합니다.", [], "")

    fields = CAPACITY_CLUSTER_FIELDS if assignment.startswith("capacity_") else EIS_CLUSTER_FIELDS
    matches = matching_condition_rows(metadata, conditions, fields)
    cluster_key = cluster_key_for(assignment, metadata, fields)
    if matches:
        status = "matched_existing_cluster"
        reason = f"{len(matches)} existing journal row(s) share the comparison fields."
    else:
        status = "new_independent_cluster"
        reason = "No existing journal row shares the comparison fields; this will start as an independent cluster."
    auto_note = ""
    if assignment == "eis_time_series" and "eis_comparison" in (item.auto_assignments or []):
        auto_note = "24hr endpoint will also be used as an EIS comparison candidate."
    return base_cluster_preview_row(item, assignment, status, reason, matches, cluster_key, auto_note=auto_note)


def base_cluster_preview_row(
    item: DraftImportFile,
    assignment: str,
    status: str,
    reason: str,
    matches: list[dict[str, object]],
    cluster_key: str,
    *,
    auto_note: str = "",
) -> dict[str, object]:
    return {
        "file_id": item.file_id,
        "filename": item.original_filename,
        "assignment": assignment,
        "assignment_label": ASSIGNMENT_LABELS.get(assignment, assignment),
        "status": status,
        "reason": reason,
        "cluster_key": cluster_key,
        "existing_match_count": len(matches),
        "existing_matches": matches[:8],
        "auto_note": auto_note,
    }


def matching_condition_rows(
    metadata: dict[str, object],
    conditions: dict[str, dict[str, object]],
    fields: tuple[str, ...],
) -> list[dict[str, object]]:
    matches = []
    for key, condition in conditions.items():
        if all(normalize_match_value(metadata.get(field)) == normalize_match_value(condition.get(field)) for field in fields):
            matches.append(
                {
                    "condition_key": key,
                    "journal_row": condition.get("_source_row_number", ""),
                    "sample": condition.get("sample") or condition.get("cell_id") or key,
                    "date": condition.get("date", ""),
                }
            )
    return sorted(matches, key=lambda row: (str(row.get("date", "")), str(row.get("sample", ""))))


def cluster_key_for(assignment: str, metadata: dict[str, object], fields: tuple[str, ...]) -> str:
    values = [assignment] + [str(metadata.get(field, "")).strip() for field in fields]
    return " | ".join(value or "-" for value in values)


def normalize_match_value(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    return text


def infer_assignment(analysis_type: str, metrics: dict[str, object], time_point: str, filename: str) -> dict[str, object]:
    lowered = f"{filename} {time_point}".lower()
    if analysis_type == "eis":
        is_time_series = bool(time_point) or re.search(r"\b(?:0|3|6|12|24)\s*hr\b", lowered) is not None
        is_24hr = "24hr" in lowered.replace(" ", "") or str(time_point).strip().lower() in {"24hr", "24 h", "24"}
        if is_time_series:
            auto = ["eis_time_series", "eis_comparison"] if is_24hr else ["eis_time_series"]
            return {
                "suggested_assignment": "eis_time_series",
                "assignment_options": list(TYPE_OPTIONS),
                "assignment_reason": "EIS time-point token detected; 24hr files are also comparison candidates." if is_24hr else "EIS time-point token detected.",
                "auto_assignments": auto,
            }
        return {
            "suggested_assignment": "eis_comparison",
            "assignment_options": list(TYPE_OPTIONS),
            "assignment_reason": "EIS file without time-series token defaults to comparison cluster.",
            "auto_assignments": ["eis_comparison"],
        }
    if analysis_type == "capacity":
        protocol = str(metrics.get("protocol") or "").upper()
        if protocol == "RATE_PERFORMANCE":
            suggested = "capacity_3"
            reason = "Capacity protocol classified as rate performance."
        elif "0P5C" in protocol or "0.5C" in protocol or protocol == "STABILIZE_THEN_0P5C":
            suggested = "capacity_2"
            reason = "Capacity protocol classified as 0.5C/stabilization."
        else:
            suggested = "capacity_1"
            reason = "Capacity protocol defaults to 0.1C continuous when no rate/stabilization signal is present."
        return {
            "suggested_assignment": suggested,
            "assignment_options": list(TYPE_OPTIONS),
            "assignment_reason": reason,
            "auto_assignments": [suggested],
        }
    return {
        "suggested_assignment": "eis_comparison",
        "assignment_options": list(TYPE_OPTIONS),
        "assignment_reason": "Analysis type unknown; defaults to EIS comparison cluster. Choose another type if needed.",
        "auto_assignments": [],
    }


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_filename(filename: str) -> str:
    name = Path(filename or "upload").name.strip()
    stem = Path(name).stem
    suffix = Path(name).suffix.lower()
    stem = re.sub(r"\s+", "_", stem)
    stem = re.sub(r"[^A-Za-z0-9_.가-힣-]+", "_", stem).strip("._")
    return f"{stem or 'upload'}{suffix}"


def unique_filename(filename: str, used_names: set[str]) -> str:
    path = Path(filename)
    stem = path.stem
    suffix = path.suffix
    candidate = filename
    index = 2
    while candidate in used_names:
        candidate = f"{stem}_{index}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate
