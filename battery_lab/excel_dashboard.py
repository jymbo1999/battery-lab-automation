from __future__ import annotations

import json
import math
import re
import threading
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

from .config import BATTERY_CONDITION_WORKBOOK

DEFAULT_CONDITION_WORKBOOK = BATTERY_CONDITION_WORKBOOK
DEFAULT_CONDITION_SHEET = "JYJ"
FILTER_RULES = {
    "참고": {"12파이_cufoil"},
    "전해질": {"1.0mlipf6ec/dec1:1"},
    "종류": {"lib"},
    "binder": {"2wt%cmc", "2wt%cmc/40wt%sbr"},
    "voltagerange": {"0.01~2v"},
}
FORMULA_TEMPLATES_BY_HEADER = {
    "current(a)": "=I{row}*S{row}*1000/10^6",
    "activematerial(g)": "=(P{row}-Q{row})*R{row}",
    "arealmassdensity(mg/cm2)": "=I{row}*1000/(PI()*(0.6)^2)",
    "전극두께(mm)": "=U{row}-V{row}",
    "electrode(g)": "=P{row}-Q{row}",
    "volume(mm3)": "=113.1*W{row}",
    "합제밀도(g/cm3)": "=X{row}/(Y{row}/1000)",
}
EXTRA_EDITABLE_ROWS = 100
FAST_VIEW_ROW_LIMIT = 40
FAST_VIEW_EXTRA_ROWS = 15
DEFAULT_VIEWER_ZOOM = 0.55

_SERVERS: dict[tuple[Path, str, str, int], "ExcelDashboardServer"] = {}


@dataclass
class ExcelDashboardServer:
    httpd: ThreadingHTTPServer
    thread: threading.Thread
    url: str


class WorkbookStore:
    def __init__(self, workbook_path: Path, sheet_name: str) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.lock = threading.Lock()

    def sheet_payload(
        self,
        include_ignored: bool = True,
        row_limit: int | None = None,
        extra_rows: int = EXTRA_EDITABLE_ROWS,
    ) -> dict[str, Any]:
        with self.lock:
            workbook = load_workbook(self.workbook_path, data_only=False)
            if self.sheet_name not in workbook.sheetnames:
                raise KeyError(f"Sheet not found: {self.sheet_name}")
            worksheet = workbook[self.sheet_name]
            payload = build_sheet_payload(
                worksheet,
                self.workbook_path,
                include_ignored=include_ignored,
                row_limit=row_limit,
                extra_rows=extra_rows,
            )
            workbook.close()
            return payload

    def update_cell(self, row: int, column: int, value: Any) -> dict[str, Any]:
        if row < 1 or column < 1:
            raise ValueError("Cell coordinates must be positive.")
        with self.lock:
            workbook = load_workbook(self.workbook_path, data_only=False)
            if self.sheet_name not in workbook.sheetnames:
                workbook.close()
                raise KeyError(f"Sheet not found: {self.sheet_name}")
            worksheet = workbook[self.sheet_name]
            cell = worksheet.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                workbook.close()
                raise ValueError("Merged child cells cannot be edited directly.")
            cell.value = normalize_edit_value(value)
            apply_row_formulas(worksheet, row)
            workbook.save(self.workbook_path)
            payload = cell_payload(cell, worksheet, self.workbook_path, formula_columns=formula_columns(worksheet), computed_cache={})
            workbook.close()
            return payload


def ensure_excel_dashboard_server(
    workbook_path: Path = DEFAULT_CONDITION_WORKBOOK,
    sheet_name: str = DEFAULT_CONDITION_SHEET,
    host: str = "127.0.0.1",
    port: int = 0,
) -> ExcelDashboardServer:
    workbook_path = workbook_path.resolve()
    key = (workbook_path, sheet_name, host, port)
    existing = _SERVERS.get(key)
    if existing and existing.thread.is_alive():
        return existing

    store = WorkbookStore(workbook_path, sheet_name)
    handler = make_handler(store)
    httpd = ThreadingHTTPServer((host, port), handler)
    actual_port = int(httpd.server_address[1])
    thread = threading.Thread(target=httpd.serve_forever, name=f"excel-dashboard-{actual_port}", daemon=True)
    thread.start()
    server = ExcelDashboardServer(httpd=httpd, thread=thread, url=f"http://{host}:{actual_port}/")
    _SERVERS[key] = server
    return server


def make_handler(store: WorkbookStore) -> type[BaseHTTPRequestHandler]:
    class Handler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:  # noqa: N802
            parsed = urlparse(self.path)
            if parsed.path == "/":
                self.send_text(render_page(), "text/html; charset=utf-8")
                return
            if parsed.path == "/api/sheet":
                try:
                    params = parse_qs(parsed.query)
                    include_ignored = (params.get("filter", ["all"])[0]).strip().lower() not in {"hide", "matched"}
                    self.send_json(
                        store.sheet_payload(
                            include_ignored=include_ignored,
                            row_limit=parse_positive_int(params.get("limit", [""])[0]),
                            extra_rows=parse_positive_int(params.get("extra", [""])[0], default=EXTRA_EDITABLE_ROWS),
                        )
                    )
                except Exception as exc:  # pragma: no cover - browser-facing error path
                    self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            self.send_error(HTTPStatus.NOT_FOUND)

        def do_POST(self) -> None:  # noqa: N802
            parsed = urlparse(self.path)
            if parsed.path != "/api/cell":
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8")
                data = json.loads(body or "{}")
                result = store.update_cell(int(data["row"]), int(data["column"]), data.get("value", ""))
                self.send_json({"ok": True, "cell": result})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, HTTPStatus.BAD_REQUEST)

        def log_message(self, format: str, *args: Any) -> None:
            return

        def send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
            encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(encoded)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(encoded)

        def send_text(self, text: str, content_type: str, status: HTTPStatus = HTTPStatus.OK) -> None:
            encoded = text.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(encoded)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(encoded)

    return Handler


def parse_positive_int(value: Any, default: int | None = None) -> int | None:
    try:
        parsed = int(str(value or "").strip())
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def build_sheet_payload(
    worksheet: Any,
    workbook_path: Path,
    include_ignored: bool = True,
    row_limit: int | None = None,
    extra_rows: int = EXTRA_EDITABLE_ROWS,
) -> dict[str, Any]:
    header_map = header_columns(worksheet)
    formula_map = formula_columns(worksheet)
    source_max_row = worksheet.max_row
    extra_rows = max(0, int(extra_rows))
    computed_cache: dict[tuple[int, int], Any] = {}
    merged_lookup: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    merged_children: set[tuple[int, int]] = set()
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_lookup[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1, max_row, max_col)
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                if (row, column) != (min_row, min_col):
                    merged_children.add((row, column))

    source_rows: list[tuple[int, bool]] = []
    ignored_count = 0
    for row_idx in range(1, source_max_row + 1):
        ignored = row_idx > 1 and not row_matches_filter(worksheet, row_idx, header_map)
        if ignored:
            ignored_count += 1
        if ignored and not include_ignored:
            continue
        source_rows.append((row_idx, ignored))

    last_body_row = max((row_idx for row_idx, _ in source_rows if row_idx > 1), default=source_max_row)
    extra_start_row = (source_max_row + 1) if include_ignored else (last_body_row + 1)
    rendered_max_row = extra_start_row + extra_rows - 1 if extra_rows else max(last_body_row, source_max_row)
    extra_template_row = last_body_row if last_body_row > 1 else 1

    partial_rows = False
    if row_limit is not None and row_limit > 0 and len(source_rows) > row_limit + 1:
        header_rows = [row for row in source_rows if row[0] == 1]
        body_rows = [row for row in source_rows if row[0] != 1]
        source_rows = header_rows + body_rows[-row_limit:]
        partial_rows = True

    row_plan = [(row_idx, row_idx, ignored, False) for row_idx, ignored in source_rows]
    row_plan.extend((row_idx, extra_template_row, False, True) for row_idx in range(extra_start_row, extra_start_row + extra_rows))

    rows = []
    for display_row_idx, source_row_idx, ignored, is_extra_row in row_plan:
        row_dimension = worksheet.row_dimensions[source_row_idx]
        row_height = points_to_px(row_dimension.height) if row_dimension.height else None
        cells = []
        for column_idx in range(1, worksheet.max_column + 1):
            if not is_extra_row and (source_row_idx, column_idx) in merged_children:
                continue
            template_cell = worksheet.cell(row=source_row_idx, column=column_idx)
            if is_extra_row:
                cells.append(extra_cell_payload(display_row_idx, column_idx, template_cell, formula_map))
            else:
                rowspan, colspan, _, _ = merged_lookup.get(
                    (source_row_idx, column_idx), (1, 1, source_row_idx, column_idx)
                )
                cells.append(
                    cell_payload(
                        template_cell,
                        worksheet,
                        workbook_path,
                        rowspan=rowspan,
                        colspan=colspan,
                        formula_columns=formula_map,
                        computed_cache=computed_cache,
                    )
                )
        rows.append(
            {
                "index": display_row_idx,
                "height": row_height,
                "hidden": bool(row_dimension.hidden),
                "ignored": ignored,
                "extra": is_extra_row,
                "cells": cells,
            }
        )

    columns = []
    for column_idx in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_idx)
        dimension = worksheet.column_dimensions[letter]
        columns.append(
            {
                "index": column_idx,
                "letter": letter,
                "width": excel_width_to_px(dimension.width),
                "hidden": bool(dimension.hidden),
            }
        )

    return {
        "title": f"실험 일지 / {worksheet.title}",
        "workbook": str(workbook_path),
        "sheet": worksheet.title,
        "maxRow": rendered_max_row,
        "sourceMaxRow": source_max_row,
        "extraRows": extra_rows,
        "extraStartRow": extra_start_row,
        "includeIgnoredRows": include_ignored,
        "partialRows": partial_rows,
        "rowLimit": row_limit,
        "renderedRowCount": len(rows),
        "maxColumn": worksheet.max_column,
        "freezePane": str(worksheet.freeze_panes or ""),
        "zoom": DEFAULT_VIEWER_ZOOM * 100,
        "filter": {
            "required": filter_description(),
            "available": all(key in header_map for key in FILTER_RULES),
            "matchedRows": max(0, source_max_row - 1 - ignored_count),
            "ignoredRows": ignored_count,
        },
        "columns": columns,
        "rows": rows,
    }


def cell_payload(
    cell: Any,
    worksheet: Any,
    workbook_path: Path,
    rowspan: int = 1,
    colspan: int = 1,
    formula_columns: dict[int, str] | None = None,
    computed_cache: dict[tuple[int, int], Any] | None = None,
) -> dict[str, Any]:
    formula_columns = formula_columns or {}
    if computed_cache is None:
        computed_cache = {}
    formula = formula_for_cell(cell.row, cell.column, formula_columns, cell.value)
    computed = evaluate_cell_value(worksheet, cell.row, cell.column, formula_columns, computed_cache) if formula else cell.value
    formula_cell = bool(formula and cell.row > 1)
    return {
        "row": cell.row,
        "column": cell.column,
        "address": cell.coordinate,
        "value": display_value(computed),
        "rawValue": display_value(cell.value),
        "formula": formula or "",
        "formulaCell": formula_cell,
        "rowspan": rowspan,
        "colspan": colspan,
        "editable": not isinstance(cell, MergedCell) and not formula_cell,
        "style": cell_style(cell),
    }


def extra_cell_payload(row: int, column: int, template_cell: Any, formula_columns: dict[int, str]) -> dict[str, Any]:
    formula = formula_for_cell(row, column, formula_columns, None)
    return {
        "row": row,
        "column": column,
        "address": f"{get_column_letter(column)}{row}",
        "value": "",
        "rawValue": "",
        "formula": formula,
        "formulaCell": bool(formula),
        "rowspan": 1,
        "colspan": 1,
        "editable": not bool(formula),
        "style": cell_style(template_cell),
    }


def formula_columns(worksheet: Any) -> dict[int, str]:
    output: dict[int, str] = {}
    for column_idx in range(1, worksheet.max_column + 1):
        key = normalize_formula_header(worksheet.cell(row=1, column=column_idx).value)
        template = FORMULA_TEMPLATES_BY_HEADER.get(key)
        if template:
            output[column_idx] = template
    return output


def normalize_formula_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    text = re.sub(r"\s+", "", text)
    return text


def formula_for_cell(row: int, column: int, formula_map: dict[int, str], raw_value: Any) -> str:
    if row <= 1:
        return str(raw_value) if isinstance(raw_value, str) and raw_value.startswith("=") else ""
    template = formula_map.get(column)
    if template:
        return template.format(row=row)
    return str(raw_value) if isinstance(raw_value, str) and raw_value.startswith("=") else ""


def apply_row_formulas(worksheet: Any, row: int) -> None:
    if row <= 1:
        return
    for column, template in formula_columns(worksheet).items():
        worksheet.cell(row=row, column=column).value = template.format(row=row)


def evaluate_cell_value(
    worksheet: Any,
    row: int,
    column: int,
    formula_map: dict[int, str],
    computed_cache: dict[tuple[int, int], Any],
    visiting: set[tuple[int, int]] | None = None,
) -> Any:
    key = (row, column)
    if key in computed_cache:
        return computed_cache[key]
    visiting = visiting or set()
    if key in visiting:
        return None
    visiting.add(key)
    cell = worksheet.cell(row=row, column=column)
    formula = formula_for_cell(row, column, formula_map, cell.value)
    if formula:
        value = evaluate_formula(worksheet, formula, formula_map, computed_cache, visiting)
    else:
        value = cell.value
    visiting.discard(key)
    computed_cache[key] = value
    return value


def evaluate_formula(
    worksheet: Any,
    formula: str,
    formula_map: dict[int, str],
    computed_cache: dict[tuple[int, int], Any],
    visiting: set[tuple[int, int]],
) -> Any:
    expression = formula[1:] if formula.startswith("=") else formula
    references: list[Any] = []

    def replace_reference(match: re.Match[str]) -> str:
        col = column_index_from_string(match.group(1))
        row = int(match.group(2))
        value = evaluate_cell_value(worksheet, row, col, formula_map, computed_cache, visiting)
        references.append(value)
        return str(float(value)) if is_numeric_value(value) else "0"

    expression = re.sub(r"\b([A-Z]+)(\d+)\b", replace_reference, expression)
    if references and all(value in (None, "") for value in references):
        return None
    expression = expression.replace("^", "**")
    expression = re.sub(r"\bPI\(\)", f"({math.pi})", expression, flags=re.IGNORECASE)
    if not re.fullmatch(r"[0-9eE+\-*/().\s]+", expression):
        return None
    try:
        result = eval(expression, {"__builtins__": {}}, {})
    except Exception:
        return None
    if isinstance(result, (int, float)) and math.isfinite(float(result)):
        return float(result)
    return None


def is_numeric_value(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def header_columns(worksheet: Any) -> dict[str, int]:
    headers = {}
    for column_idx in range(1, worksheet.max_column + 1):
        value = normalize_filter_value(worksheet.cell(row=1, column=column_idx).value)
        if value:
            headers[value] = column_idx
    return headers


def row_matches_filter(worksheet: Any, row_idx: int, header_map: dict[str, int]) -> bool:
    for header, allowed_values in FILTER_RULES.items():
        column_idx = header_map.get(header)
        if column_idx is None:
            return True
        value = normalize_filter_value(worksheet.cell(row=row_idx, column=column_idx).value)
        if value not in allowed_values:
            return False
    return True


def normalize_filter_value(value: Any) -> str:
    text = str(value or "").strip().lower()
    return "".join(text.split())


def filter_description() -> list[str]:
    return [
        "참고 = 12파이_Cu foil",
        "전해질 = 1.0M LiPF6 EC/DEC 1:1",
        "종류 = LIB",
        "Binder = 2wt% cmc 또는 2wt%cmc/40wt%SBR",
        "Voltage range = 0.01~2V",
    ]


def normalize_edit_value(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text == "":
        return None
    if text.startswith("="):
        return text
    try:
        if "." not in text and "e" not in text.lower():
            return int(text)
        return float(text)
    except ValueError:
        return value


def cell_style(cell: Any) -> dict[str, Any]:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    return {
        "fontFamily": font.name,
        "fontSize": points_to_px(font.sz) if font.sz else None,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "underline": bool(font.underline),
        "color": openpyxl_color(font.color),
        "backgroundColor": openpyxl_color(fill.fgColor) if fill and fill.fill_type else None,
        "horizontal": css_horizontal_alignment(alignment.horizontal),
        "vertical": css_vertical_alignment(alignment.vertical),
        "wrapText": bool(alignment.wrap_text),
        "border": border_style(border),
    }


def openpyxl_color(color: Any) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        if len(rgb) == 8:
            alpha = int(rgb[:2], 16)
            if alpha == 0:
                return None
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    return None


def css_horizontal_alignment(value: str | None) -> str | None:
    if value in {None, "general", "fill", "centerContinuous", "distributed", "justify"}:
        return None if value in {None, "general"} else "center"
    return value


def css_vertical_alignment(value: str | None) -> str | None:
    mapping = {
        "top": "top",
        "center": "middle",
        "bottom": "bottom",
        "justify": "middle",
        "distributed": "middle",
    }
    return mapping.get(value or "")


def border_style(border: Any) -> dict[str, str]:
    return {
        side: border_color(getattr(border, side))
        for side in ("left", "right", "top", "bottom")
        if getattr(border, side).style
    }


def border_color(side: Any) -> str:
    color = openpyxl_color(side.color) or "#d0d7de"
    return f"1px solid {color}"


def points_to_px(points: float) -> int:
    return max(1, round(float(points) * 96 / 72))


def excel_width_to_px(width: float | None) -> int:
    if width is None:
        width = 8.43
    return max(24, round(float(width) * 7 + 5))


def render_page(
    sheet_api_url: str = "/api/sheet",
    cell_api_url: str = "/api/cell",
    row_types_api_url: str = "",
    row_detail_api_url: str = "",
) -> str:
    page = """<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>실험 일지</title>
  <style>
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: #f3f4f6;
      color: #111827;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Apple SD Gothic Neo", "Noto Sans KR", sans-serif;
      letter-spacing: 0;
    }
    .bar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      min-height: 52px;
      padding: 10px 14px;
      background: #ffffff;
      border-bottom: 1px solid #d0d7de;
      position: sticky;
      top: 0;
      z-index: 5;
    }
    .title { min-width: 0; }
    .title strong { display: block; font-size: 15px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .title span { display: block; color: #6b7280; font-size: 12px; margin-top: 2px; }
    .rule-note {
      border-bottom: 1px solid #d0d7de;
      background: #fbfcfe;
      color: #6b7280;
      font-size: 11px;
      line-height: 1.45;
      padding: 5px 14px 6px;
    }
    .tools {
      display: flex;
      align-items: center;
      gap: 12px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }
    .filter-control {
      display: flex;
      align-items: center;
      gap: 7px;
      color: #57606a;
      font-size: 13px;
      white-space: nowrap;
    }
    .filter-control select {
      min-width: 146px;
      border: 1px solid #d0d7de;
      border-radius: 6px;
      background: #ffffff;
      color: #111827;
      padding: 6px 8px;
      outline: none;
    }
    .zoom-control {
      display: flex;
      align-items: center;
      gap: 6px;
      color: #57606a;
      font-size: 13px;
      white-space: nowrap;
    }
    .zoom-control button {
      width: 30px;
      height: 30px;
      border: 1px solid #d0d7de;
      border-radius: 6px;
      background: #ffffff;
      color: #111827;
      cursor: pointer;
      font-size: 17px;
      line-height: 1;
    }
    .zoom-control output {
      min-width: 46px;
      text-align: center;
      color: #57606a;
      font-variant-numeric: tabular-nums;
    }
    .status { color: #6b7280; font-size: 13px; white-space: nowrap; }
    .sheet-wrap {
      height: calc(100vh - 79px);
      overflow: auto;
      background: #f8fafc;
    }
    table.sheet {
      border-collapse: separate;
      border-spacing: 0;
      table-layout: fixed;
      background: #ffffff;
      font-size: 12px;
    }
    col.row-head { width: 48px; }
    thead th {
      position: sticky;
      top: 0;
      z-index: 3;
      height: 26px;
      background: #f6f8fa;
      color: #57606a;
      border-right: 1px solid #d0d7de;
      border-bottom: 1px solid #d0d7de;
      font-weight: 600;
      text-align: center;
    }
    thead th.corner {
      left: 0;
      z-index: 4;
    }
    tbody th {
      position: sticky;
      left: 0;
      z-index: 2;
      min-width: 48px;
      background: #f6f8fa;
      color: #57606a;
      border-right: 1px solid #d0d7de;
      border-bottom: 1px solid #d0d7de;
      font-weight: 600;
      text-align: center;
    }
    tbody tr.header-row th,
    tbody tr.header-row td {
      position: sticky;
      top: 26px;
      z-index: 3;
      box-shadow: 0 1px 0 #d0d7de;
    }
    tbody tr.header-row th {
      z-index: 5;
    }
    td {
      min-height: 22px;
      padding: 4px 6px;
      border-right: 1px solid #d0d7de;
      border-bottom: 1px solid #d0d7de;
      vertical-align: middle;
      overflow: hidden;
      white-space: pre-wrap;
      overflow-wrap: anywhere;
      line-height: 1.25;
      background: #ffffff;
      outline: none;
      cursor: cell;
      user-select: none;
    }
    td:focus {
      box-shadow: inset 0 0 0 2px #1a73e8;
      z-index: 1;
      position: relative;
    }
    td.selected-cell {
      background: #dbeafe !important;
      box-shadow: inset 0 0 0 1px #2563eb;
    }
    td.selection-anchor {
      box-shadow: inset 0 0 0 2px #1d4ed8;
    }
    td.formula-cell {
      color: #374151;
      cursor: default;
      box-shadow: inset 3px 0 0 #1a73e8;
    }
    tbody tr.ignored th,
    td.ignored {
      background: #e5e7eb !important;
      color: #6b7280 !important;
    }
    tbody th.row-head-cell { cursor: pointer; }
    tbody th.row-head-cell:hover { background: #eaeef2; color: #1a73e8; }
    tbody th.row-head-cell.row-no-data {
      background: #9ca3af;
      color: #1f2937;
    }
    tbody th.row-head-cell.row-no-data:hover { background: #8b919b; color: #111827; }
    tbody th.row-head-cell.row-active {
      background: #dbeafe;
      color: #1d4ed8;
      box-shadow: inset 0 0 0 2px #2563eb;
    }
    tbody th.row-head-cell.row-no-data.row-active { background: #93a3c9; }
    tbody tr.row-selected td { background: #eef4ff; }
    .row-popup-backdrop {
      position: fixed;
      inset: 0;
      background: rgba(15, 23, 42, 0.45);
      display: none;
      align-items: flex-start;
      justify-content: center;
      padding: 28px 16px;
      overflow: auto;
      z-index: 50;
    }
    .row-popup-backdrop.open { display: flex; }
    .row-popup {
      width: min(1040px, 100%);
      background: #fff;
      border-radius: 12px;
      box-shadow: 0 18px 48px rgba(15, 23, 42, 0.28);
      overflow: hidden;
    }
    .row-popup-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 14px 18px;
      border-bottom: 1px solid #e5e9f0;
    }
    .row-popup-head h2 { margin: 0; font-size: 16px; }
    .row-popup-head .sub { color: #6b7280; font-size: 12px; margin-top: 2px; }
    .row-popup-close {
      width: 32px; height: 32px; border: none; border-radius: 7px;
      background: #eef1f5; color: #374151; font-size: 19px; line-height: 1; cursor: pointer;
    }
    .row-popup-close:hover { background: #e2e8f0; }
    .row-popup-body { padding: 16px 18px; max-height: calc(100vh - 160px); overflow: auto; }
    .rp-section-title { font-size: 13px; font-weight: 700; color: #334155; margin: 6px 0 8px; }
    .rp-preview { display: grid; grid-template-columns: repeat(auto-fit, minmax(360px, 1fr)); gap: 12px; margin-bottom: 14px; }
    .rp-box { border: 1px solid #d8dee8; border-radius: 10px; background: #fff; padding: 10px; overflow: hidden; }
    .rp-box-head { display: flex; align-items: center; gap: 8px; justify-content: space-between; margin-bottom: 6px; }
    .rp-chip { font-size: 11px; padding: 2px 8px; border-radius: 6px; background: #eef2f7; color: #475569; }
    .rp-frame { width: 100%; height: 420px; border: 1px solid #e5e9f0; border-radius: 6px; background: #fff; }
    .rp-info-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 10px; }
    .rp-field { display: flex; flex-direction: column; gap: 4px; font-size: 12px; color: #475569; }
    .rp-field input { min-height: 32px; border: 1px solid #a8b3c2; border-radius: 7px; padding: 5px 8px; font-size: 13px; }
    .rp-field input:disabled { background: #f1f5f9; color: #94a3b8; }
    .rp-actions { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 14px; align-items: center; }
    .rp-btn { display: inline-flex; align-items: center; min-height: 34px; border: 1px solid #a8b3c2; border-radius: 7px; padding: 6px 12px; background: #fff; color: #1f2937; font-size: 13px; cursor: pointer; }
    .rp-btn.primary { background: #2563eb; border-color: #2563eb; color: #fff; }
    .rp-btn:disabled { opacity: 0.5; cursor: not-allowed; }
    .rp-msg { font-size: 12px; color: #475569; margin-left: auto; }
    .rp-empty { color: #647084; padding: 18px; background: #f8fafc; border: 1px dashed #bdc7d5; border-radius: 8px; font-size: 13px; }
    .rp-replace { border: 1px solid #fed7aa; background: #fff7ed; border-radius: 10px; padding: 12px; margin-bottom: 16px; }
    .rp-replace .rp-section-title { margin-top: 0; }
    .rp-replace-row { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
    .rp-replace-select { flex: 1 1 220px; min-height: 34px; border: 1px solid #a8b3c2; border-radius: 7px; background: #fff; padding: 5px 8px; font-size: 12px; }
    .rp-replace-input { flex: 1 1 200px; font-size: 12px; }
    .rp-replace-raw { display: inline-flex; align-items: center; gap: 5px; font-size: 12px; color: #475569; white-space: nowrap; }
    .rp-replace-msg { font-size: 12px; color: #9a3412; margin-top: 8px; min-height: 16px; }
    .hidden { display: none; }
  </style>
</head>
<body>
  <div class="bar">
    <div class="title">
      <strong id="title">실험 일지</strong>
      <span id="meta">JYJ 시트를 불러오는 중입니다.</span>
    </div>
    <div class="tools">
      <label class="filter-control" for="filterMode">
        필터
        <select id="filterMode">
          <option value="hide" selected>무시 행 표시안함</option>
          <option value="gray">무시 행 회색</option>
          <option value="all">전체 보기</option>
        </select>
      </label>
      <div class="zoom-control" aria-label="뷰어 배율">
        <button type="button" id="zoomOut" title="축소">-</button>
        <output id="zoomValue">55%</output>
        <button type="button" id="zoomIn" title="확대">+</button>
      </div>
      <div class="status" id="status">Loading</div>
    </div>
  </div>
  <div class="rule-note" id="ruleNote">
    필터 기준: 참고=12파이_Cu foil · 전해질=1.0M LiPF6 EC/DEC 1:1 · 종류=LIB · Binder=2wt% cmc 또는 2wt%cmc/40wt%SBR · Voltage range=0.01~2V
  </div>
  <div class="sheet-wrap" id="sheet"></div>
  <div class="row-popup-backdrop" id="rowPopup">
    <div class="row-popup" role="dialog" aria-modal="true">
      <div class="row-popup-head">
        <div>
          <h2 id="rowPopupTitle">행 미리보기</h2>
          <div class="sub" id="rowPopupSub"></div>
        </div>
        <button type="button" class="row-popup-close" id="rowPopupClose" aria-label="닫기">&times;</button>
      </div>
      <div class="row-popup-body" id="rowPopupBody"></div>
    </div>
  </div>
  <script>
    const api = {
      sheet: __SHEET_API_URL__,
      cell: __CELL_API_URL__,
      rowTypes: __ROW_TYPES_API_URL__,
      rowDetail: __ROW_DETAIL_API_URL__,
    };
    const TYPE_LABELS = {
      eis: 'EIS',
      eis_time_series: 'EIS time series',
      eis_comparison: 'EIS',
      type_1_0p1c_continuous: 'capacity 1',
      type_2_0p5c_after_stabilization: 'capacity 2',
      type_3_rate_performance: 'capacity 3',
      capacity_1: 'capacity 1',
      capacity_2: 'capacity 2',
      capacity_3: 'capacity 3',
      capacity: 'capacity',
    };
    const state = { data: null, filterMode: 'hide', zoom: 0.55, didInitialScroll: false, selection: null, dragging: false, fullDataLoaded: false, loadingFilteredRows: false, selectedRowIndex: null, rowTypes: {}, orphanRows: new Set() };
    const statusEl = document.getElementById('status');
    const titleEl = document.getElementById('title');
    const metaEl = document.getElementById('meta');
    const sheetEl = document.getElementById('sheet');
    const filterModeEl = document.getElementById('filterMode');
    const zoomOutEl = document.getElementById('zoomOut');
    const zoomInEl = document.getElementById('zoomIn');
    const zoomValueEl = document.getElementById('zoomValue');
    filterModeEl.addEventListener('change', () => {
      state.filterMode = filterModeEl.value;
      if (!state.data) return;
      if (state.filterMode !== 'hide' && !state.fullDataLoaded) {
        loadSheet(true, { preserveScroll: true }).catch(error => {
          sheetEl.textContent = error.message;
          setStatus('Error');
        });
        return;
      }
      renderSheet(state.data);
    });
    zoomOutEl.addEventListener('click', () => setZoom(state.zoom - 0.08, viewportCenterAnchor()));
    zoomInEl.addEventListener('click', () => setZoom(state.zoom + 0.08, viewportCenterAnchor()));
    sheetEl.addEventListener('wheel', event => {
      if (!event.ctrlKey && !event.metaKey) return;
      event.preventDefault();
      const step = event.deltaY > 0 ? -0.06 : 0.06;
      setZoom(state.zoom + step, { clientX: event.clientX, clientY: event.clientY });
    }, { passive: false });
    document.addEventListener('pointerup', () => { state.dragging = false; });
    document.addEventListener('copy', event => {
      const copied = copySelectedCells(event);
      if (copied) setStatus(`Copied ${copied.rows}x${copied.columns}`);
    });

    function setStatus(text) { statusEl.textContent = text; }
    function viewportCenterAnchor() {
      const rect = sheetEl.getBoundingClientRect();
      return { clientX: rect.left + rect.width / 2, clientY: rect.top + rect.height / 2 };
    }
    function setZoom(value, anchor = null) {
      const oldZoom = state.zoom;
      const nextZoom = Math.max(0.45, Math.min(1.8, value));
      const table = sheetEl.querySelector('table.sheet');
      let anchorPoint = null;
      if (table && anchor && oldZoom > 0) {
        const rect = sheetEl.getBoundingClientRect();
        const viewportX = anchor.clientX - rect.left;
        const viewportY = anchor.clientY - rect.top;
        anchorPoint = {
          viewportX,
          viewportY,
          contentX: (sheetEl.scrollLeft + viewportX) / oldZoom,
          contentY: (sheetEl.scrollTop + viewportY) / oldZoom,
        };
      }
      state.zoom = nextZoom;
      zoomValueEl.textContent = `${Math.round(state.zoom * 100)}%`;
      if (table) table.style.zoom = state.zoom;
      if (anchorPoint) {
        sheetEl.scrollLeft = anchorPoint.contentX * state.zoom - anchorPoint.viewportX;
        sheetEl.scrollTop = anchorPoint.contentY * state.zoom - anchorPoint.viewportY;
      }
    }

    function sheetUrl(includeIgnored, fastView = false) {
      const separator = api.sheet.includes('?') ? '&' : '?';
      const params = new URLSearchParams({ filter: includeIgnored ? 'all' : 'hide' });
      if (fastView && !includeIgnored) {
        params.set('limit', '40');
        params.set('extra', '15');
      }
      return `${api.sheet}${separator}${params.toString()}`;
    }

    async function loadSheet(includeIgnored = false, options = {}) {
      setStatus('Loading');
      const response = await fetch(sheetUrl(includeIgnored, Boolean(options.fastView)));
      const data = await response.json();
      if (data.error) throw new Error(data.error);
      state.data = data;
      state.fullDataLoaded = Boolean(data.includeIgnoredRows);
      titleEl.textContent = data.title;
      const filter = data.filter || {};
      const filterText = filter.available === false
        ? '필터 기준 헤더를 찾지 못했습니다.'
        : `조건 만족 ${filter.matchedRows ?? 0}행 · 무시 ${filter.ignoredRows ?? 0}행`;
      const partialText = data.partialRows ? ` · 빠른 보기 ${data.renderedRowCount ?? data.rows.length}행` : '';
      metaEl.textContent = `${data.maxRow} rows x ${data.maxColumn} columns · ${filterText}${partialText}`;
      renderSheet(data);
      if (!options.preserveScroll) scrollToLatestRows();
      setStatus('Ready');
      if (data.partialRows && options.fastView) loadCompleteFilteredRows();
    }

    async function loadCompleteFilteredRows() {
      if (state.loadingFilteredRows || state.fullDataLoaded) return;
      state.loadingFilteredRows = true;
      setStatus('Loading all matched rows');
      try {
        const response = await fetch(sheetUrl(false, false));
        const data = await response.json();
        if (data.error) throw new Error(data.error);
        if (state.fullDataLoaded || state.filterMode !== 'hide') {
          setStatus('Ready');
          return;
        }
        state.data = data;
        titleEl.textContent = data.title;
        const filter = data.filter || {};
        const filterText = filter.available === false
          ? '필터 기준 헤더를 찾지 못했습니다.'
          : `조건 만족 ${filter.matchedRows ?? 0}행 · 무시 ${filter.ignoredRows ?? 0}행`;
        metaEl.textContent = `${data.maxRow} rows x ${data.maxColumn} columns · ${filterText}`;
        if (state.filterMode === 'hide') {
          renderSheet(data);
          scrollToLatestRows(true);
        }
        setStatus('Ready');
      } catch (error) {
        setStatus(`Partial view: ${error.message}`);
      } finally {
        state.loadingFilteredRows = false;
      }
    }

    function scrollToLatestRows(force = false) {
      if (state.didInitialScroll && !force) return;
      state.didInitialScroll = true;
      requestAnimationFrame(() => {
        const extraStartRow = Number(state.data?.extraStartRow || 0);
        const sourceMaxRow = Number(state.data?.sourceMaxRow || state.data?.maxRow || 1);
        const targetIndex = extraStartRow || sourceMaxRow;
        const targetRow = sheetEl.querySelector(`tr[data-row-index="${targetIndex}"]`)
          || sheetEl.querySelector(`tr[data-row-index="${sourceMaxRow}"]`);
        if (!targetRow) {
          sheetEl.scrollTop = sheetEl.scrollHeight;
          return;
        }
        const visibleExtraRows = 6;
        const rowHeight = Math.max(1, targetRow.getBoundingClientRect().height || targetRow.offsetHeight * state.zoom);
        const targetTop = targetRow.offsetTop * state.zoom;
        sheetEl.scrollTop = Math.max(0, targetTop - sheetEl.clientHeight + rowHeight * visibleExtraRows);
      });
    }

    function renderSheet(data) {
      const table = document.createElement('table');
      table.className = 'sheet';
      table.style.zoom = state.zoom;
      const colgroup = document.createElement('colgroup');
      const rowHead = document.createElement('col');
      rowHead.className = 'row-head';
      colgroup.appendChild(rowHead);
      data.columns.forEach(column => {
        const col = document.createElement('col');
        col.style.width = `${column.width}px`;
        if (column.hidden) col.className = 'hidden';
        colgroup.appendChild(col);
      });
      table.appendChild(colgroup);

      const thead = document.createElement('thead');
      const headRow = document.createElement('tr');
      const corner = document.createElement('th');
      corner.className = 'corner';
      headRow.appendChild(corner);
      data.columns.forEach(column => {
        const th = document.createElement('th');
        th.textContent = column.letter;
        if (column.hidden) th.className = 'hidden';
        headRow.appendChild(th);
      });
      thead.appendChild(headRow);
      table.appendChild(thead);

      const tbody = document.createElement('tbody');
      data.rows.forEach(row => {
        if (row.ignored && state.filterMode === 'hide') return;
        const tr = document.createElement('tr');
        tr.dataset.rowIndex = row.index;
        if (row.height) tr.style.height = `${row.height}px`;
        if (row.hidden) tr.className = 'hidden';
        if (row.index === 1) tr.classList.add('header-row');
        if (row.ignored && state.filterMode === 'gray') tr.classList.add('ignored');
        const rowLabel = document.createElement('th');
        rowLabel.textContent = row.index;
        if (row.index !== 1) {
          rowLabel.classList.add('row-head-cell');
          rowLabel.dataset.rowIndex = row.index;
          applyRowTooltip(rowLabel, row.index);
          rowLabel.addEventListener('click', () => handleRowHeadClick(row.index));
          rowLabel.addEventListener('dblclick', event => { event.preventDefault(); openRowPopup(row.index); });
        }
        tr.appendChild(rowLabel);
        row.cells.forEach(cell => {
          const td = document.createElement('td');
          td.dataset.row = cell.row;
          td.dataset.column = cell.column;
          td.dataset.original = cell.value ?? '';
          td.dataset.formula = cell.formula ?? '';
          td.textContent = cell.value ?? '';
          td.rowSpan = cell.rowspan || 1;
          td.colSpan = cell.colspan || 1;
          td.tabIndex = 0;
          td.contentEditable = 'false';
          applyStyle(td, cell.style || {});
          if (cell.formulaCell) {
            td.classList.add('formula-cell');
            td.title = cell.formula || '';
          }
          if (row.ignored && state.filterMode === 'gray') td.classList.add('ignored');
          td.addEventListener('pointerdown', event => {
            if (event.button !== 0) return;
            event.preventDefault();
            state.dragging = true;
            state.selectedRowIndex = null;
            selectCells(cell.row, cell.column, cell.row, cell.column);
            td.focus({ preventScroll: true });
          });
          td.addEventListener('pointerenter', () => {
            if (!state.dragging || !state.selection) return;
            selectCells(state.selection.anchorRow, state.selection.anchorColumn, cell.row, cell.column);
          });
          td.addEventListener('keydown', event => {
            if ((event.metaKey || event.ctrlKey) && event.key.toLowerCase() === 'c') return;
            if (event.key === 'Escape') clearSelection();
          });
          tr.appendChild(td);
        });
        tbody.appendChild(tr);
      });
      table.appendChild(tbody);
      sheetEl.replaceChildren(table);
      setZoom(state.zoom);
      applySelection();
    }

    function selectCells(anchorRow, anchorColumn, focusRow, focusColumn) {
      state.selection = { anchorRow, anchorColumn, focusRow, focusColumn };
      applySelection();
    }

    function clearSelection() {
      state.selection = null;
      state.selectedRowIndex = null;
      applySelection();
    }

    function selectWholeRow(rowIndex) {
      const maxColumn = Number(state.data?.maxColumn || 1);
      state.selectedRowIndex = rowIndex;
      selectCells(rowIndex, 1, rowIndex, maxColumn);
    }

    function handleRowHeadClick(rowIndex) {
      // Second click on an already-selected row opens the detail popup.
      if (state.selectedRowIndex === rowIndex) {
        openRowPopup(rowIndex);
        return;
      }
      selectWholeRow(rowIndex);
    }

    function applyRowHeadState() {
      sheetEl.querySelectorAll('tbody tr').forEach(tr => {
        const head = tr.querySelector('th.row-head-cell');
        const isActive = head && Number(head.dataset.rowIndex) === state.selectedRowIndex;
        tr.classList.toggle('row-selected', Boolean(isActive));
        if (head) head.classList.toggle('row-active', Boolean(isActive));
      });
    }

    function typeLabelsForRow(rowIndex) {
      const types = state.rowTypes[rowIndex] || state.rowTypes[String(rowIndex)] || [];
      return types.map(t => TYPE_LABELS[t] || t);
    }

    function applyRowTooltip(rowLabel, rowIndex) {
      const labels = typeLabelsForRow(rowIndex);
      const isOrphan = state.orphanRows.has(rowIndex);
      rowLabel.title = labels.length
        ? `데이터 유형: ${labels.join(', ')}`
        : (isOrphan ? '데이터 파일 없음' : '데이터 정보를 불러오는 중...');
      rowLabel.classList.toggle('row-no-data', isOrphan && !labels.length);
    }

    function refreshRowTooltips() {
      sheetEl.querySelectorAll('th.row-head-cell').forEach(head => {
        applyRowTooltip(head, Number(head.dataset.rowIndex));
      });
    }

    async function loadRowTypes() {
      if (!api.rowTypes) return;
      try {
        const response = await fetch(api.rowTypes, { headers: { Accept: 'application/json' } });
        if (!response.ok) return;
        const payload = await response.json();
        state.rowTypes = (payload && payload.row_types) || {};
        state.orphanRows = new Set(((payload && payload.orphan_rows) || []).map(Number));
        refreshRowTooltips();
      } catch (error) { /* tooltip is best-effort */ }
    }

    function selectionBounds() {
      if (!state.selection) return null;
      return {
        top: Math.min(state.selection.anchorRow, state.selection.focusRow),
        bottom: Math.max(state.selection.anchorRow, state.selection.focusRow),
        left: Math.min(state.selection.anchorColumn, state.selection.focusColumn),
        right: Math.max(state.selection.anchorColumn, state.selection.focusColumn),
      };
    }

    function applySelection() {
      const bounds = selectionBounds();
      sheetEl.querySelectorAll('td[data-row][data-column]').forEach(td => {
        const row = Number(td.dataset.row);
        const column = Number(td.dataset.column);
        const selected = bounds && row >= bounds.top && row <= bounds.bottom && column >= bounds.left && column <= bounds.right;
        td.classList.toggle('selected-cell', Boolean(selected));
        td.classList.toggle('selection-anchor', Boolean(state.selection && row === state.selection.anchorRow && column === state.selection.anchorColumn));
      });
      applyRowHeadState();
    }

    function selectedCellMatrix() {
      const bounds = selectionBounds();
      if (!bounds) return null;
      const byAddress = new Map();
      sheetEl.querySelectorAll('td[data-row][data-column]').forEach(td => {
        byAddress.set(`${td.dataset.row}:${td.dataset.column}`, td.dataset.original || td.textContent || '');
      });
      const rows = [];
      for (let row = bounds.top; row <= bounds.bottom; row += 1) {
        const values = [];
        for (let column = bounds.left; column <= bounds.right; column += 1) {
          values.push(byAddress.get(`${row}:${column}`) || '');
        }
        rows.push(values);
      }
      return rows;
    }

    function copySelectedCells(event) {
      const rows = selectedCellMatrix();
      if (!rows) return null;
      const text = rows.map(row => row.join('\\t')).join('\\n');
      event.preventDefault();
      event.clipboardData.setData('text/plain', text);
      return { rows: rows.length, columns: rows[0]?.length || 0 };
    }

    function applyStyle(td, style) {
      if (style.fontFamily) td.style.fontFamily = `"${style.fontFamily}", sans-serif`;
      if (style.fontSize) td.style.fontSize = `${style.fontSize}px`;
      if (style.bold) td.style.fontWeight = '700';
      if (style.italic) td.style.fontStyle = 'italic';
      if (style.underline) td.style.textDecoration = 'underline';
      if (style.color) td.style.color = style.color;
      if (style.backgroundColor) td.style.backgroundColor = style.backgroundColor;
      if (style.horizontal) td.style.textAlign = style.horizontal;
      if (style.vertical) td.style.verticalAlign = style.vertical;
      if (style.wrapText === false) td.style.whiteSpace = 'nowrap';
      const border = style.border || {};
      Object.entries(border).forEach(([side, value]) => { td.style[`border${side[0].toUpperCase()}${side.slice(1)}`] = value; });
    }

    const popupEl = document.getElementById('rowPopup');
    const popupTitleEl = document.getElementById('rowPopupTitle');
    const popupSubEl = document.getElementById('rowPopupSub');
    const popupBodyEl = document.getElementById('rowPopupBody');
    const popupCloseEl = document.getElementById('rowPopupClose');
    let popupRow = null;

    function closeRowPopup() {
      popupEl.classList.remove('open');
      popupBodyEl.innerHTML = '';
      popupRow = null;
    }
    popupCloseEl.addEventListener('click', closeRowPopup);
    popupEl.addEventListener('click', event => { if (event.target === popupEl) closeRowPopup(); });
    document.addEventListener('keydown', event => { if (event.key === 'Escape' && popupEl.classList.contains('open')) closeRowPopup(); });

    async function openRowPopup(rowIndex) {
      popupRow = rowIndex;
      popupTitleEl.textContent = `행 ${rowIndex} · 데이터 미리보기 & 실험정보`;
      const labels = typeLabelsForRow(rowIndex);
      popupSubEl.textContent = labels.length ? `데이터 유형: ${labels.join(', ')}` : '등록된 데이터 유형 확인 중...';
      popupBodyEl.innerHTML = '<div class="rp-empty">불러오는 중...</div>';
      popupEl.classList.add('open');
      if (!api.rowDetail) {
        popupBodyEl.innerHTML = '<div class="rp-empty">상세 보기 API가 구성되지 않았습니다.</div>';
        return;
      }
      try {
        const url = `${api.rowDetail}${api.rowDetail.includes('?') ? '&' : '?'}row=${encodeURIComponent(rowIndex)}`;
        const response = await fetch(url, { headers: { Accept: 'application/json' } });
        const payload = await response.json();
        if (!response.ok || payload.error) throw new Error(payload.error || '행 정보를 불러오지 못했습니다.');
        renderRowPopup(payload);
      } catch (error) {
        popupBodyEl.innerHTML = `<div class="rp-empty">${escapeText(error.message)}</div>`;
      }
    }

    function escapeText(value) {
      const div = document.createElement('div');
      div.textContent = String(value ?? '');
      return div.innerHTML;
    }

    function renderRowPopup(payload) {
      if (popupRow !== payload.row) return;
      const labels = (payload.types || []).map(t => TYPE_LABELS[t] || t);
      popupSubEl.textContent = labels.length ? `데이터 유형: ${labels.join(', ')}` : '이 행에 매칭된 데이터 파일이 없습니다.';

      const previews = payload.previews || [];

      // --- Replace section (top): pick which linked file to swap, upload, recompute. ---
      const replaceHtml = (payload.replace_url && previews.length)
        ? `<div class="rp-replace">
            <div class="rp-section-title">데이터 파일 교체 <span style="font-weight:400;color:#94a3b8;font-size:11px">— 새 파일로 덮어쓰고(.bak 백업) 지표·클러스터를 완전히 재계산합니다. 실험정보는 그대로 유지됩니다.</span></div>
            <div class="rp-replace-row">
              <select class="rp-replace-select" data-role="replace-target">
                ${previews.map(p => `<option value="${escapeText(p.file)}" data-kind="${escapeText(p.kind || '')}">${escapeText((TYPE_LABELS[p.type] || p.type || '') + ' · ' + (p.title || p.file))}</option>`).join('')}
              </select>
              <input type="file" class="rp-replace-input" data-role="replace-input" accept=".sde,.seo,.wrd,.csv,.xlsx,.xls">
              <label class="rp-replace-raw" title="WRD 교체 시 원시 time-series CSV도 생성"><input type="checkbox" data-role="replace-write-raw"> 원시 CSV도</label>
              <button type="button" class="rp-btn primary" data-role="do-replace">교체 & 재계산</button>
            </div>
            <div class="rp-replace-msg" data-role="replace-msg"></div>
          </div>`
        : (previews.length ? '' : '<div class="rp-empty">이 행에 연결된 데이터 파일이 없어 교체할 대상이 없습니다.</div>');

      const previewHtml = previews.length
        ? `<div class="rp-preview">${previews.map((p, i) => `
            <div class="rp-box">
              <div class="rp-box-head">
                <strong style="font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${escapeText(p.title || p.file || '미리보기')}</strong>
                <span class="rp-chip">${escapeText(TYPE_LABELS[p.type] || p.type || '')}</span>
              </div>
              ${p.available === false
                ? `<div class="rp-empty">${escapeText((p.errors && p.errors[0]) || '미리보기를 만들 수 없습니다.')}</div>`
                : `<iframe class="rp-frame" data-preview-index="${i}" sandbox="allow-scripts" title="${escapeText(p.title || 'preview')}"></iframe>`}
            </div>`).join('')}</div>`
        : '<div class="rp-empty">이 행에 연결된 데이터 파일이 없습니다.</div>';

      const fields = payload.info_fields || [];
      const infoHtml = fields.length
        ? `<div class="rp-info-grid">${fields.map(f => `
            <label class="rp-field">
              ${escapeText(f.header || f.letter)}
              <input type="text" data-row="${f.row}" data-column="${f.column}" data-original="${escapeText(f.value ?? '')}" value="${escapeText(f.value ?? '')}" ${f.editable === false ? 'disabled' : ''}>
            </label>`).join('')}</div>`
        : '<div class="rp-empty">표시할 실험정보 칸이 없습니다.</div>';

      popupBodyEl.innerHTML = `
        ${replaceHtml}
        <div class="rp-section-title">미리보기</div>
        ${previewHtml}
        <div class="rp-section-title">실험정보 <span style="font-weight:400;color:#94a3b8;font-size:11px">— 값을 고치고 ‘수정내용 저장하기’를 누르면 실험일지에 반영됩니다.</span></div>
        ${infoHtml}
        <div class="rp-actions">
          <button type="button" class="rp-btn primary" data-role="save-info" ${fields.length ? '' : 'disabled'}>수정내용 저장하기</button>
          <span class="rp-msg" data-role="rp-msg"></span>
        </div>`;

      previews.forEach((p, i) => {
        if (p.available === false || !p.html) return;
        const frame = popupBodyEl.querySelector(`iframe[data-preview-index="${i}"]`);
        if (frame) frame.srcdoc = p.html;
      });

      const msgEl = popupBodyEl.querySelector('[data-role="rp-msg"]');
      const saveBtn = popupBodyEl.querySelector('[data-role="save-info"]');
      if (saveBtn) saveBtn.addEventListener('click', () => saveRowInfo(payload.row, msgEl));
      const replaceBtn = popupBodyEl.querySelector('[data-role="do-replace"]');
      if (replaceBtn) replaceBtn.addEventListener('click', () => replaceRowFile(payload));
    }

    async function saveRowInfo(rowIndex, msgEl) {
      const inputs = Array.from(popupBodyEl.querySelectorAll('.rp-info-grid input:not(:disabled)'));
      const changed = inputs.filter(input => input.value !== (input.dataset.original ?? ''));
      if (!changed.length) {
        if (msgEl) msgEl.textContent = '변경된 칸이 없습니다.';
        return;
      }
      if (msgEl) msgEl.textContent = '저장 중...';
      let failures = 0;
      for (const input of changed) {
        try {
          const response = await fetch(api.cell, {
            method: 'POST',
            headers: { 'Content-Type': 'application/json' },
            body: JSON.stringify({ row: Number(input.dataset.row), column: Number(input.dataset.column), value: input.value }),
          });
          const payload = await response.json();
          if (!response.ok || !payload.ok) throw new Error(payload.error || 'save failed');
          input.dataset.original = input.value;
        } catch (error) { failures += 1; }
      }
      if (msgEl) msgEl.textContent = failures ? `${changed.length - failures}칸 저장, ${failures}칸 실패` : `${changed.length}칸 저장 완료`;
      // Reload the underlying sheet so formula columns reflect the new values.
      loadSheet(state.filterMode !== 'hide', { preserveScroll: true }).catch(() => {});
    }

    async function replaceRowFile(payload) {
      const select = popupBodyEl.querySelector('[data-role="replace-target"]');
      const fileInput = popupBodyEl.querySelector('[data-role="replace-input"]');
      const writeRaw = popupBodyEl.querySelector('[data-role="replace-write-raw"]');
      const msgEl = popupBodyEl.querySelector('[data-role="replace-msg"]');
      const button = popupBodyEl.querySelector('[data-role="do-replace"]');
      if (!select || !fileInput || !fileInput.files || !fileInput.files.length) {
        if (msgEl) msgEl.textContent = '교체할 새 파일을 선택하세요.';
        return;
      }
      const target = select.value;
      const kind = select.selectedOptions[0]?.dataset.kind || '';
      if (!window.confirm(`선택한 파일을 새 파일로 교체하고 지표·클러스터를 재계산합니다.\n기존 파일은 .bak로 백업됩니다.\n\n대상: ${target}\n\n진행할까요?`)) return;
      const form = new FormData();
      form.append('row', String(payload.row));
      form.append('kind', kind);
      form.append('target', target);
      form.append('file', fileInput.files[0]);
      if (writeRaw && writeRaw.checked) form.append('write_raw', '1');
      if (button) button.disabled = true;
      if (msgEl) msgEl.textContent = '교체 및 재계산 중... (수십 초 걸릴 수 있습니다)';
      try {
        const response = await fetch(payload.replace_url, { method: 'POST', body: form });
        const result = await response.json();
        if (!response.ok || result.ok === false) throw new Error(result.error || '교체에 실패했습니다.');
        const failed = (result.recompute || []).filter(r => r.ok === false);
        if (msgEl) msgEl.textContent = failed.length
          ? `교체 완료 · 재계산 일부 실패: ${failed.map(f => f.kind).join(', ')}`
          : `교체 & 재계산 완료 → ${result.new_rel_path}`;
        // Refresh tooltips/types and re-open the row to show the new preview.
        await loadRowTypes();
        loadSheet(state.filterMode !== 'hide', { preserveScroll: true }).catch(() => {});
        openRowPopup(payload.row);
      } catch (error) {
        if (msgEl) msgEl.textContent = `오류: ${error.message}`;
        if (button) button.disabled = false;
      }
    }

    loadSheet(false, { fastView: true }).then(() => loadRowTypes()).catch(error => {
      sheetEl.textContent = error.message;
      setStatus('Error');
    });
  </script>
</body>
</html>
"""
    return (
        page.replace("__SHEET_API_URL__", json.dumps(sheet_api_url))
        .replace("__CELL_API_URL__", json.dumps(cell_api_url))
        .replace("__ROW_TYPES_API_URL__", json.dumps(row_types_api_url))
        .replace("__ROW_DETAIL_API_URL__", json.dumps(row_detail_api_url))
    )
